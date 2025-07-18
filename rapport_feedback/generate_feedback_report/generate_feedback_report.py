import calendar

import numpy as np
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import utils
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.defined_name import DefinedName

# Formule générale qui sera formatée pour les différentes règles
f_rule_etat_stock = 'NOT(ISERROR(SEARCH("{etat_stock}", {col_letter_etat_stock}{index_start})))'


def remove_empty_rows(workbook, sheet_name):
    """
    Supprime les lignes vides de la feuille spécifiée dans le classeur.
    Paramètres:
        workbook (openpyxl.Workbook): Le classeur contenant la feuille.
        sheet_name (str): Le nom de la feuille dont il faut supprimer les lignes vides.
    Exception:
        AssertionError: Si sheet_name n'est pas dans le classeur.
    """
    assert sheet_name in workbook.sheetnames, f"{sheet_name} not in sheetnames"

    sheet = workbook[sheet_name]
    rows_to_delete = [
        row[0].row for row in sheet.iter_rows() if all(cell.value is None for cell in row)
    ]

    for row_idx in reversed(rows_to_delete):
        sheet.delete_rows(row_idx)


def export_detail_comp_promp_to_sheet(
    wb_feedback_report, df_ets: pd.DataFrame, df_region: pd.DataFrame, date_report
):
    """
    Cette fonction est utilisée principalement pour exporter les données vers les feuilles:
        - DetailCompletude
        - DetailPromptitude
        - CompletudeParRegion
        - PromptitudeParRegion
    """
    date_report = pd.to_datetime(date_report)

    ws_comp, ws_promp = (
        wb_feedback_report["DetailCompletude"],
        wb_feedback_report["DetailPromptitude"],
    )

    ws_region_comp, ws_region_promp = (
        wb_feedback_report["CompletudeParRegion"],
        wb_feedback_report["PromptitudeParRegion"],
    )

    cols = [
        col
        for col in [
            "ARV",
            "TRC",
            "LAB",
            "CHARGE VIRALE",
            "PNLP",
            "PNSME-GRAT",
            "PNSME",
            "PNN",
            "PNLT",
            "TBS",
            "TBMR",
            "TBLAB",
        ]
        if col in df_ets.columns
    ]

    df_ets["Attendu"] = df_ets[cols].apply(
        lambda row: len([element for element in row if element != "NA"]), axis=1
    )

    df_ets["Reçu"] = df_ets[cols].apply(
        lambda row: len([element for element in row if element == 1]), axis=1
    )

    df_comp, df_promp = (
        df_ets.loc[df_ets["Indicateur type"] == "Completude"],
        df_ets.loc[df_ets["Indicateur type"] == "Promptitude"],
    )
    df_region_comp, df_region_promp = (
        df_region.loc[
            (df_region["Indicateur type"] == "Completude") & (df_region.Region != "NATIONAL")
        ],
        df_region.loc[
            (df_region["Indicateur type"] == "Promptitude") & (df_region.Region != "NATIONAL")
        ],
    )

    header_row = list(ws_comp.iter_rows(min_row=1, max_row=1))[0]
    list_cols = list(df_comp.columns)
    dico_cols = {
        cell.value.rstrip(): [
            cell.column_letter,
            cell.col_idx,
            list_cols.index(cell.value.rstrip().replace("\n", " ")),
        ]
        for cell in header_row
        if cell.value is not None and cell.value.rstrip().replace("\n", " ") in list_cols
    }

    # Mise en forme des cellules
    font = Font(name="Arial Narrow", size=11)
    alignment_one = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_two = Alignment(
        horizontal=None, vertical="center", wrap_text=True
    )  # Pour la Region & Site

    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, bottom=side)
    start = 2
    # for start, row_comp, row_promp in enumerate(dataframe_to_rows(df_comp, index=False, header=False), dataframe_to_rows(df_promp, index=False, header=False), start=2):
    for row_comp, row_promp in zip(
        dataframe_to_rows(df_comp, index=False, header=False),
        dataframe_to_rows(df_promp, index=False, header=False),
    ):
        for col, element in dico_cols.items():
            cell_comp = ws_comp.cell(row=start, column=element[1], value=row_comp[element[2]])
            cell_promp = ws_promp.cell(row=start, column=element[1], value=row_promp[element[2]])
            # Gestion de la mise en forme des cellules
            cell_comp.font = cell_promp.font = font
            cell_comp.alignment = cell_promp.alignment = (
                alignment_one if col not in ("Site", "Region") else alignment_two
            )
            cell_comp.border = cell_promp.border = border

        start += 1

    fill = PatternFill(start_color="FFFFC000", fill_type="solid")
    font = Font(name="Arial Narrow", size=14, bold=True)
    for col in cols:
        if date_report.month not in (3, 6, 9, 12) and col in ("PNLT", "TBS", "TBMR", "TBLAB"):
            value_comp = value_promp = "NA"
            cell_comp = ws_comp.cell(row=start, column=dico_cols[col][1], value=value_comp)
            cell_promp = ws_promp.cell(row=start, column=dico_cols[col][1], value=value_promp)
        else:
            value_comp, value_promp = (
                df_comp.loc[df_comp[col] != "NA", col].astype(int).mean(),
                df_promp.loc[df_promp[col] != "NA", col].astype(int).mean(),
            )
            cell_comp = ws_comp.cell(row=start, column=dico_cols[col][1], value=value_comp)
            cell_promp = ws_promp.cell(row=start, column=dico_cols[col][1], value=value_promp)
            cell_comp.number_format = cell_promp.number_format = numbers.FORMAT_PERCENTAGE

        cell_comp.font = cell_promp.font = font
        cell_comp.fill = cell_promp.fill = fill
        cell_comp.alignment = cell_promp.alignment = alignment_one

    # df_comp['sum_produit_inline'].sum()/ df_comp['count_produit_inline'].sum()
    # Gestion de la dernière ligne
    col_start, col_end = dico_cols[cols[0]][0], dico_cols[cols[-1]][0]
    ws_comp.merge_cells(f"{col_start}{start + 1}:{col_end}{start + 1}")
    ws_promp.merge_cells(f"{col_start}{start + 1}:{col_end}{start + 1}")

    cell_comp = ws_comp.cell(
        row=start + 1,
        column=dico_cols[cols[0]][1],
        value=df_comp["sum_produit_inline"].sum() / df_comp["count_produit_inline"].sum(),
    )
    cell_promp = ws_promp.cell(
        row=start + 1,
        column=dico_cols[cols[0]][1],
        value=df_promp["sum_produit_inline"].sum() / df_promp["count_produit_inline"].sum(),
    )

    cell_comp.font = cell_promp.font = Font(name="Arial Narrow", size=20, bold=True)
    cell_comp.fill = cell_promp.fill = PatternFill(start_color="FF00B0F0", fill_type="solid")
    cell_comp.alignment = cell_promp.alignment = alignment_one
    cell_comp.number_format = cell_promp.number_format = numbers.FORMAT_PERCENTAGE
    ws_comp.row_dimensions[start + 1].height = ws_promp.row_dimensions[start + 1].height = 35

    # Conditionnal formatting
    rule_na = CellIsRule(
        operator="equal",
        formula=['"NA"'],
        fill=PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"),
        font=Font(name="Arial Narrow", size=11),
    )

    rule_zero = CellIsRule(
        operator="equal",
        formula=["0"],
        fill=PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
        font=Font(name="Arial Narrow", size=11),
    )

    rule_one = CellIsRule(
        operator="equal",
        formula=["1"],
        fill=PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid"),
        font=Font(name="Arial Narrow", size=11, bold=True),
    )

    # ------> Mise en forme conditionnelle produits et programme
    # --------------> Ajout sur la Feuille détaile complétude
    ws_comp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_na)
    ws_comp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_zero)
    ws_comp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_one)

    # --------------> Ajout sur la Feuille détaile promptitude
    ws_promp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_na)
    ws_promp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_zero)
    ws_promp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_one)

    # ------> Mise en forme conditionnelle PNLS reçu et PNLS attendu
    col_start, col_end = dico_cols["PNLS\nrecu"][0], dico_cols["PNLS\nattendu"][0]

    # --------------> Ajout de la mise en forme conditionnelle
    ws_comp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_na)
    ws_comp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_zero)
    ws_comp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_one)

    # --------------> Ajout de la mise en forme conditionnelle
    ws_promp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_na)
    ws_promp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_zero)
    ws_promp.conditional_formatting.add(f"{col_start}2:{col_end}{start - 1}", rule_one)

    # Taux par Région des différents produits et programmes
    df_count_region = (
        df_comp.groupby("Region")["Code"]
        .count()
        .reset_index()
        .sort_values(["Region"], ascending=True)
    )
    df_count_region["Index_cols"] = None

    # Initialisation de la valeur pour "ABIDJAN 1" et "ABIDJAN 2"
    df_count_region.loc[0, "Index_cols"] = 2
    df_count_region.loc[1, "Index_cols"] = df_count_region.loc[0, "Code"] + 2

    # Définition des index
    for i in range(2, len(df_count_region)):
        df_count_region.loc[i, "Index_cols"] = (
            df_count_region.loc[i - 1, "Index_cols"] + df_count_region.loc[i - 1, "Code"]
        )

    df_count_region["Index_for_merging_cell"] = df_count_region[["Code", "Index_cols"]].apply(
        lambda row: "{0}" + str(row[1]) + ":{0}" + str(row[0] + row[1] - 1), axis=1
    )
    df_region_comp = df_region_comp[
        ["Region"] + [col for col in df_region_comp.columns if "Taux par Région" in col]
    ].replace(0.0, np.nan)
    df_region_promp = df_region_promp[
        ["Region"] + [col for col in df_region_promp.columns if "Taux par Région" in col]
    ].replace(0.0, np.nan)

    # Merging with df_count_region
    df_region_comp = df_region_comp.merge(
        df_count_region[["Region", "Index_cols", "Index_for_merging_cell"]], how="left"
    )  # .drop(columns='Region')
    df_region_promp = df_region_promp.merge(
        df_count_region[["Region", "Index_cols", "Index_for_merging_cell"]], how="left"
    )  # .drop(columns='Region')
    del df_count_region

    # Jointure des cellules taux par région en fonction du nombre de région
    header_row = list(
        ws_comp.iter_rows(
            max_row=1, min_col=dico_cols[cols[-1]][1] + 1, max_col=dico_cols["PNLS\nrecu"][1] - 1
        )
    )[0]
    list_cols = list(df_region_comp.columns)
    dico_cols = {
        cell.value: [
            cell.column_letter,
            cell.col_idx,
            list_cols.index(cell.value.rstrip().replace("\n", " ")),
        ]
        for cell in header_row
    }

    # Mise en forme des éléméents de taux par région
    side = Side(style="medium", color="000000")
    border = Border(right=side, top=side, bottom=side)

    font = Font(name="Arial Narrow", size=24, bold=True)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill = PatternFill(start_color="FFEDEDED", fill_type="solid")

    # Affectation des valeurs aux cellules
    start = 2
    for index_col_merge, index_cols in zip(
        df_region_comp["Index_for_merging_cell"].values, df_region_comp["Index_cols"].values
    ):
        for col, element in dico_cols.items():
            ws_comp.merge_cells(
                index_col_merge.format(element[0])
            )  # Petit problème ici à revoir pourquoi il ne merge pas
            ws_promp.merge_cells(index_col_merge.format(element[0]))

            for row in ws_comp[index_col_merge.format(element[0])]:
                for cell in row:
                    cell.border = border

            for row in ws_promp[index_col_merge.format(element[0])]:
                for cell in row:
                    cell.border = border

            ws_comp.merge_cells(
                index_col_merge.format(element[0])
            )  # Petit problème ici à revoir pourquoi il ne merge pas
            ws_promp.merge_cells(index_col_merge.format(element[0]))

            # Affectation des valeurs
            value_comp = df_region_comp.iloc[start - 2, dico_cols[col][2]]
            cell_comp = ws_comp.cell(row=index_cols, column=element[1], value=value_comp)
            value_promp = df_region_promp.iloc[start - 2, dico_cols[col][2]]
            cell_promp = ws_promp.cell(row=index_cols, column=element[1], value=value_promp)
            cell_comp.font = cell_promp.font = font
            cell_comp.fill = cell_promp.fill = fill
            cell_comp.alignment = cell_promp.alignment = alignment
            cell_comp.number_format = cell_promp.number_format = numbers.FORMAT_PERCENTAGE

        start += 1

    # Gestion des feuilles ComplétudeParRegion & PromptitudeRegion
    df_region_comp, df_region_promp = (
        df_region.loc[
            (df_region["Indicateur type"] == "Completude") & (df_region.Region != "NATIONAL")
        ],
        df_region.loc[
            (df_region["Indicateur type"] == "Promptitude") & (df_region.Region != "NATIONAL")
        ],
    )
    df_region_comp = (
        df_region_comp.rename(
            columns={
                "Total rapports attendus CHARGE VIRALE": "Total rapports attendus Charges Virales",
                "Taux par Région Charges virales": "Taux par Région Charges Virales",
            }
        )
        .drop(columns=["Indicateur type", "taux_indicateur_region"])
        .rename(
            columns=lambda x: x.replace("par Région", "de Completude") if "par Région" in x else x
        )
    )

    df_region_promp = (
        df_region_promp.rename(
            columns={
                "Total rapports attendus CHARGE VIRALE": "Total rapports attendus Charges Virales",
                "Taux par Région Charges virales": "Taux par Région Charges Virales",
            }
        )
        .drop(columns=["Indicateur type", "taux_indicateur_region"])
        .rename(
            columns=lambda x: x.replace("par Région", "de Promptitude") if "par Région" in x else x
        )
    )

    header_row_first = [
        cell.value
        for cell in list(ws_region_comp.iter_rows(min_row=3, max_row=3, min_col=2))[0]
        if cell.value is not None
    ]  # ws_region_comp
    header_row = []
    i = 2
    for col in header_row_first:
        header_row.append(["Total rapports attendus" + " " + col, i])
        header_row.append(["Taux de Completude" + " " + col, i + 1])
        i += 2
    del i, col, header_row_first
    list_cols = list(df_region_comp.columns)
    dico_cols = {val[0]: [val[1], list_cols.index(val[0])] for val in header_row}
    dico_cols["Region"] = [1, 0]
    font = Font(name="Arial Narrow", size=12)
    font_region = Font(name="Arial Narrow", size=12, bold=True)

    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_region = Alignment(horizontal=None, vertical="center", wrap_text=True)

    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, bottom=side)
    border_two = Border(left=side, right=Side(style="medium", color="000000"), bottom=side)

    fill = PatternFill(start_color="FFFFFFFF", fill_type="solid")
    fill_comp = PatternFill(start_color="FFDDEBF7", fill_type="solid")
    fill_promp = PatternFill(start_color="FFE2EFDA", fill_type="solid")

    start = 5
    for row_comp, row_promp in zip(
        dataframe_to_rows(df_region_comp, index=False, header=False),
        dataframe_to_rows(df_region_promp, index=False, header=False),
    ):
        for col, element in dico_cols.items():
            # Affectation des valeurs
            val_comp, val_promp = row_comp[element[1]], row_promp[element[1]]
            cell_comp = ws_region_comp.cell(row=start, column=element[0], value=val_comp)
            cell_promp = ws_region_promp.cell(row=start, column=element[0], value=val_promp)
            # print(col, val_comp, val_promp, start, element[0])
            # Gestion de la mise en forme des cellules
            cell_comp.font = cell_promp.font = font if col != "Region" else font_region
            cell_comp.alignment = cell_promp.alignment = (
                alignment if col != "Region" else alignment_region
            )
            cell_comp.fill = fill_comp if start % 2 == 0 else fill
            cell_promp.fill = fill_promp if start % 2 == 0 else fill
            if "Taux de Completude" in col:
                cell_comp.number_format = cell_promp.number_format = numbers.FORMAT_PERCENTAGE
                cell_comp.border = cell_promp.border = border_two

            if "Total rapports attendus" in col:
                cell_comp.number_format = cell_promp.number_format = "#,##0"
                cell_comp.border = cell_promp.border = border

        start += 1

    # Gestion des deux dernières lignes
    fill_comp = PatternFill(start_color="FFFFC000", fill_type="solid")
    fill_promp = PatternFill(start_color="FFC6E0B4", fill_type="solid")
    side = Side(style="medium", color="000000")
    border = Border(left=side, right=side, bottom=side, top=side)

    prog = ["pnlp", "pnsme", "pnn", "pnlt", "pnls"]
    ws_sheet_two = wb_feedback_report["Feuil2"]  # Gestion de certains éléments de la feuille 2
    start_prog = 41
    # print(df_region_comp.columns)
    for col, element in dico_cols.items():
        # Affectation des valeurs
        if col == "Region":
            cell_comp = ws_region_comp.cell(row=start, column=element[0], value="GLOBAL")
            cell_promp = ws_region_promp.cell(row=start, column=element[0], value="GLOBAL")
            cell_comp.font = cell_comp.font = Font(name="Arial Narrow", size=12, bold=True)
            cell_comp.alignment = cell_promp.alignment = alignment

        if "Total rapports attendus" in col:
            cell_comp = ws_region_comp.cell(
                row=start, column=element[0], value=df_region_comp[col].sum()
            )
            cell_promp = ws_region_promp.cell(
                row=start, column=element[0], value=df_region_promp[col].sum()
            )
            cell_comp.number_format = cell_promp.number_format = "#,##0"
            cell_comp.font = cell_comp.font = Font(name="Arial Narrow", size=14, bold=True)
            cell_comp.alignment = cell_promp.alignment = alignment
            cell_comp.fill, cell_promp.fill = fill_comp, fill_promp

        if ("Taux de Completude" in col) and col.split()[-1].lower() not in prog:
            _ = col.split(" ")[-1]
            _ = _ if _ != "Virales" else "CHARGE VIRALE"
            value_comp = df_comp.loc[df_comp[_] != "NA", _].astype(int).mean()
            value_promp = df_promp.loc[df_promp[_] != "NA", _].astype(int).mean()
            cell_comp = ws_region_comp.cell(row=start, column=element[0], value=value_comp)
            cell_promp = ws_region_promp.cell(row=start, column=element[0], value=value_promp)
            cell_comp.font = cell_comp.font = Font(name="Arial Narrow", size=14, bold=True)
            cell_comp.number_format = cell_promp.number_format = numbers.FORMAT_PERCENTAGE
            cell_comp.alignment = cell_promp.alignment = alignment
            cell_comp.border = cell_promp.border = border
            cell_comp.fill, cell_promp.fill = fill_comp, fill_promp

        if ("Total rapports attendus" not in col) and (col.split()[-1].lower() in prog):
            value_comp = df_comp["taux_indicateur_" + col.split()[-1].lower()].unique()[0]
            value_promp = df_promp["taux_indicateur_" + col.split()[-1].lower()].unique()[0]

            cell_comp = ws_region_comp.cell(row=start, column=element[0], value=value_comp)
            cell_promp = ws_region_promp.cell(row=start, column=element[0], value=value_promp)
            cell_comp.font = cell_comp.font = Font(name="Arial Narrow", size=14, bold=True)
            cell_comp.number_format = cell_promp.number_format = numbers.FORMAT_PERCENTAGE
            cell_comp.alignment = cell_promp.alignment = alignment
            cell_comp.border = cell_promp.border = border
            cell_comp.fill, cell_promp.fill = fill_comp, fill_promp

            # Affectation des éléments de la complétude sur la feuille 2
            if col.split()[-1].lower() == "pnls":
                ws_sheet_two["H40"] = df_comp[
                    "taux_indicateur_" + col.split()[-1].lower()
                ].unique()[0]
                ws_sheet_two["H40"].number_format = numbers.FORMAT_PERCENTAGE
            else:
                ws_sheet_two[f"H{start_prog}"] = df_comp[
                    "taux_indicateur_" + col.split()[-1].lower()
                ].unique()[0]
                ws_sheet_two[f"H{start_prog}"].number_format = numbers.FORMAT_PERCENTAGE
                start_prog += 1

    # Gestion de la dernière ligne
    del dico_cols["Region"], start_prog
    cols = list(dico_cols)
    col_start, col_end = (
        utils.get_column_letter(dico_cols[cols[0]][0]),
        utils.get_column_letter(dico_cols[cols[-1]][0]),
    )
    ws_region_comp.merge_cells(f"{col_start}{start + 1}:{col_end}{start + 1}")
    ws_region_promp.merge_cells(f"{col_start}{start + 1}:{col_end}{start + 1}")

    cell_comp = ws_region_comp.cell(
        row=start + 1,
        column=dico_cols[cols[0]][0],
        value=df_comp["sum_produit_inline"].sum() / df_comp["count_produit_inline"].sum(),
    )
    cell_promp = ws_region_promp.cell(
        row=start + 1,
        column=dico_cols[cols[0]][0],
        value=df_promp["sum_produit_inline"].sum() / df_promp["count_produit_inline"].sum(),
    )

    cell_comp.font = cell_promp.font = Font(name="Arial Narrow", size=20, bold=True)
    cell_comp.fill = PatternFill(start_color="FF0070C0", fill_type="solid")
    cell_promp.fill = PatternFill(start_color="FF548235", fill_type="solid")
    cell_comp.alignment = cell_promp.alignment = alignment
    cell_comp.number_format = cell_promp.number_format = numbers.FORMAT_PERCENTAGE
    ws_region_comp.row_dimensions[start + 1].height = ws_region_promp.row_dimensions[
        start + 1
    ].height = 30


def export_stock_data_to_sheet(
    wb_feedback_report, extract_stock: pd.DataFrame, dest_file: str = ""
):
    """
    Cette fonction permet de formatter les données de la feuille ETAT DU STOCK du template
    """

    # wb_feedback_report = load_workbook(filename=src_file)

    ws = wb_feedback_report["ETAT DU STOCK"]

    header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]

    extract_stock = extract_stock.rename(
        columns={"BESOIN CMMMANDE URGENTE": "BESOIN COMMANDE URGENTE"}
    ).drop(columns=["id_region_esigl", "id_district_esigl"])

    extract_stock["Code_Pro"] = extract_stock["CODE"].astype(str) + "_" + extract_stock["PROGRAMME"]

    cols_extract_stock = list(extract_stock.columns)
    dico_cols = {}
    for cell in header_row:
        _ = cell.value.replace("  ", " ").rstrip()
        if _ in cols_extract_stock:
            dico_cols[cell.value.rstrip()] = [
                cell.column_letter,
                cell.col_idx,
                cols_extract_stock.index(_),
                max([len(val) for val in extract_stock[_].astype(str)]),
            ]

    # Gestion de la mise en forme des données
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, bottom=side)
    font_one = Font(name="Arial Narrow", size=11, bold=True)
    font_two = Font(name="Arial Narrow", size=11)
    fill = PatternFill(start_color="FFDDEBF7", fill_type="solid")
    alignment_one = Alignment(horizontal=None, vertical="center", wrap_text=True)
    alignment_two = Alignment(horizontal="center", vertical="center", wrap_text=True)

    special_font_cols = {
        "CODE",
        "CODE  ETS",
        "CMM gestionnaire",
        "QUANTITE PROPOSEE",
        "MSD",
        "ETAT DU STOCK",
        "BESOIN COMMANDE URGENTE",
        "BESOIN TRANSFERT IN",
        "QUANTITE A TRANSFERER OUT",
    }

    special_alignment_cols = {"Code_Pro", "CODE", "PROGRAMME", "CODE  ETS", "CATEGORIE PRODUIT"}
    special_alignment_indices = range(14, 30)

    number_format_cols = {
        "CMM gestionnaire": "#,##0",
        "Code_Pro": "0.00",
        "CATEGORIE PRODUIT": "0.00",
        "MSD": "0.00",
    }
    fill_cols = {
        "Code_Pro",
        "PROGRAMME",
        "CATEGORIE PRODUIT",
        "CMM gestionnaire",
        "MSD",
        "ETAT DU STOCK",
        "BESOIN COMMANDE URGENTE",
        "BESOIN TRANSFERT IN",
        "QUANTITE A TRANSFERER OUT",
    }

    for start, row in enumerate(
        dataframe_to_rows(extract_stock, index=False, header=False), start=2
    ):
        ws.row_dimensions[start].height = 25.2
        for col, element in dico_cols.items():
            cell = ws.cell(row=start, column=element[1], value=row[element[2]])
            cell.border = border
            cell.font = font_one if col in special_font_cols else font_two
            cell.alignment = (
                alignment_two
                if col in special_alignment_cols or element[1] in special_alignment_indices
                else alignment_one
            )
            if col in number_format_cols:
                cell.number_format = number_format_cols[col]
            if col in fill_cols:
                cell.fill = fill
        cell = ws.cell(row=start, column=element[1] + 1)
        cell.border = border
        cell.font = font_one
        cell = ws.cell(row=start, column=element[1] + 2)
        cell.border = border
        cell.font = font_one

    # Add rule in column Categorie du produit
    rule_categorie_produit = CellIsRule(
        operator="equal",
        formula=['"Produit traceur"'],
        fill=PatternFill(start_color="FF8EA9DB", end_color="FF8EA9DB", fill_type="solid"),
        font=Font(name="Arial Narrow", size=11, bold=True),
    )

    ws.conditional_formatting.add(
        f"{dico_cols['CATEGORIE PRODUIT'][0]}2:{dico_cols['CATEGORIE PRODUIT'][0]}{start}",
        rule_categorie_produit,
    )

    # Add rule in column Etat du stock
    col_letter_etat_stock = dico_cols["ETAT DU STOCK"][0]

    dico_dxf = {
        "STOCK DORMANT": DifferentialStyle(
            fill=PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True, color="FFFFFFFF"),
        ),
        "SURSTOCK": DifferentialStyle(
            fill=PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True),
        ),
        "RUPTURE": DifferentialStyle(
            fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True, color="FFFFFFFF"),
        ),
        "BIEN STOCKE": DifferentialStyle(
            fill=PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True),
        ),
        "EN BAS DU PCU": DifferentialStyle(
            fill=PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True),
        ),
        "ENTRE PCU et MIN": DifferentialStyle(
            fill=PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True),
        ),
    }
    for etat_stock in dico_dxf:
        rule = Rule(
            type="containsText",
            operator="containsText",
            dxf=dico_dxf[etat_stock],
            text=etat_stock,
            formula=[
                f_rule_etat_stock.format(
                    etat_stock=etat_stock,
                    col_letter_etat_stock=col_letter_etat_stock,
                    index_start=2,
                )
            ],
        )

        ws.conditional_formatting.add(
            f"{col_letter_etat_stock}2:{col_letter_etat_stock}{start}", rule
        )

    # Gestion des largeurs des colonnes
    # for element in dico_cols.values():
    #     ws.column_dimensions[element[0]].width = (element[3]+5)*1.2

    formula = (
        utils.quote_sheetname("ETAT DU STOCK")
        + "!"
        + f"${utils.get_column_letter(ws.min_column)}$1:${utils.get_column_letter(ws.max_column)}${start}"
    )

    wb_feedback_report.defined_names["ETAT_DU_STOCK"] = DefinedName(
        name="ETAT_DU_STOCK", attr_text=formula, localSheetId=6
    )

    remove_empty_rows(wb_feedback_report, "ETAT DU STOCK")

    if dest_file:
        wb_feedback_report.save(dest_file)


def export_stock_region_to_sheet(
    wb_feedback_report,
    stock_lvl_decent: pd.DataFrame,
    stock_region: pd.DataFrame,
    dest_file: str = "",
    date_report="2024/04/11",
):
    """
    Cette fonction permet de formatter les données de la feuille StockParRegion et certains éléments de la feuille2 du template
    """
    df_ = stock_lvl_decent.copy()

    df_["Code"] = df_["Code"].astype(str) + "_" + df_["Programme"]

    df_two = stock_region.copy()
    df_two["Code"] = df_two["Code"].astype(str) + "_" + df_two["Programme"]

    for region in df_two["Region"].drop_duplicates().sort_values().values:
        df_[f"{region}_MSD"] = df_.merge(
            df_two.loc[df_two.Region == region], how="left", on="Code"
        )["MSD"].fillna("NA")

        df_[f"{region}_STATUT"] = df_.merge(
            df_two.loc[df_two.Region == region], how="left", on="Code"
        )["STATUT"].fillna("NA")

    df_.rename(
        columns={
            "Designation": "Désignation",
            "Unite": "Unité",
            "lvl_decent_conso": "CONSO",
            "lvl_decent_sdu": "SDU",
            "lvl_decent_cmm": "CMM",
            "lvl_decent_msd": "MSD",
            "lvl_decent_statut": "STATUT",
        },
        inplace=True,
    )

    ws = wb_feedback_report["StockParRegion"]
    ws_sheet_two = wb_feedback_report["Feuil2"]

    list_cols = list(df_.columns)
    header_row = list(ws.iter_rows(min_row=2, max_row=2))[0]
    dico_cols = {
        cell.value.rstrip(): [
            cell.column_letter,
            cell.col_idx,
            list_cols.index(cell.value.rstrip()),
        ]
        for cell in header_row
        if cell.value is not None and cell.value in list_cols
    }

    header_row = list(ws.iter_rows(min_row=3, max_row=3, min_col=6, max_col=10))[0]

    dico_cols.update(
        {
            cell.value: [cell.column_letter, cell.col_idx, list_cols.index(cell.value.rstrip())]
            for cell in header_row
            if cell.value is not None and cell.value in list_cols
        }
    )

    header_row = list(ws.iter_rows(min_row=2, max_row=2, min_col=11))[0]
    header_row = [
        (cell.value, cell.column_letter, cell.col_idx)
        for cell in header_row
        if cell.value is not None
    ]
    for element in header_row:
        if f"{element[0]}_MSD" in list_cols:
            dico_cols[f"{element[0]}_MSD"] = [
                element[1],
                element[2],
                list_cols.index(f"{element[0]}_MSD"),
            ]
            dico_cols[f"{element[0]}_STATUT"] = [
                utils.get_column_letter(element[2] + 1),
                element[2] + 1,
                list_cols.index(f"{element[0]}_STATUT"),
            ]

    dico_cols["Categorie_produit"] = ["BY", 77, list_cols.index("Categorie_produit")]

    dico_cols_two = {"Programme": ["T", 20], "Désignation": ["U", 21]}

    del header_row, list_cols

    # Gestion de la mise en forme des données
    side = Side(style="thin", color="000000")
    border_one = Border(left=side, right=side, bottom=side)
    font_one = Font(name="Arial Narrow", size=10, bold=True)
    alignment_one = Alignment(horizontal="center", vertical="center", wrap_text=True)

    special_font_cols = {"Désignation", "Categorie", "Unité", "Categorie_produit"}
    font_two = Font(name="Arial Narrow", size=10)

    special_alignment_cols = {"Code", "Désignation", "Categorie", "Unité"}
    alignment_two = Alignment(horizontal=None, vertical="center", wrap_text=True)

    special_border_cols = {"Unité", "STATUT"}
    border_two = Border(left=side, right=Side(style="medium", color="000000"), bottom=side)

    number_format_cols = {"CONSO": "#,##0", "SDU": "#,##0", "CMM": "#,##0"}

    fill_cols = {"Code", "Programme", "STATUT"}
    fill = PatternFill(start_color="FFD9D9D9", fill_type="solid")

    for start, row in enumerate(dataframe_to_rows(df_, index=False, header=False), start=4):
        ws.row_dimensions[start].height = 36
        for col, element in dico_cols.items():
            cell = ws.cell(row=start, column=element[1], value=row[element[2]])
            # Par la suite je dois refaire la mise en forme du document également
            cell.border = border_one if col not in special_border_cols else border_two
            cell.font = (
                font_one
                if (col not in special_font_cols) or ("MSD" in col) or ("STATUT" in col)
                else font_two
            )
            cell.alignment = (
                alignment_one
                if (col not in special_alignment_cols) or ("MSD" in col) or ("STATUT" in col)
                else alignment_two
            )
            if col in number_format_cols:
                cell.number_format = number_format_cols[col]
            if "MSD" in col:
                cell.number_format = "0.00"
            if col in fill_cols:
                cell.fill = fill

            # Gestion des éléments de la feuille 2
            if col in dico_cols_two:
                cell_two = ws_sheet_two.cell(
                    row=start - 2, column=dico_cols_two[col][1], value=row[element[2]]
                )
                cell_two.font = font_one if col == "Programme" else font_two
                cell_two.alignment = alignment_one if col == "Programme" else alignment_two
                cell_two.border = border_one
                if col == "Programme":
                    cell_two.fill = fill

    remove_empty_rows(wb_feedback_report, "StockParRegion")
    dico_dxf = {
        "STOCK DORMANT": DifferentialStyle(
            fill=PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True, color="FFFFFFFF"),
        ),
        "SURSTOCK": DifferentialStyle(
            fill=PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid"),
            font=Font(name="Arial Narrow", size=10, bold=True),
        ),
        "RUPTURE": DifferentialStyle(
            fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
            font=Font(name="Arial Narrow", size=10, bold=True, color="FFFFFFFF"),
        ),
        "BIEN STOCKE": DifferentialStyle(
            fill=PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid"),
            font=Font(name="Arial Narrow", size=10, bold=True),
        ),
        "SOUS-STOCK": DifferentialStyle(
            fill=PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),
            font=Font(name="Arial Narrow", size=10, bold=True),
        ),
        "NA": DifferentialStyle(
            fill=PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"),
            font=Font(name="Arial Narrow", size=10, bold=True),
        ),
    }
    # Gestion des règles
    for col in dico_cols:
        if "STATUT" in col:
            for etat_stock in dico_dxf:
                col_letter_etat_stock = dico_cols[col][0]
                rule = Rule(
                    type="containsText",
                    operator="containsText",
                    dxf=dico_dxf[etat_stock],
                    text=etat_stock,
                    formula=[
                        f_rule_etat_stock.format(
                            etat_stock=etat_stock,
                            col_letter_etat_stock=col_letter_etat_stock,
                            index_start=4,
                        )
                    ],
                )
                ws.conditional_formatting.add(
                    f"{col_letter_etat_stock}4:{col_letter_etat_stock}{start}", rule
                )

    dico_prog_prod = {
        prog: df_.loc[df_.Programme == prog, "Désignation"].values
        for prog in ["PNLS", "PNLT", "PNLP", "PNSME", "PNN"]
    }
    min_row = 2
    for prog, designations in dico_prog_prod.items():
        for col_offset, value in enumerate(designations, start=25):
            ws_sheet_two.cell(row=min_row, column=col_offset, value=value)
        min_row += 1

    # Modification de la date visualisée dans le volet acceuil
    ws_acceuil = wb_feedback_report["Accueil"]
    dico_mois_fr = {
        "January": "Janvier",
        "February": "Février",
        "March": "Mars",
        "April": "Avril",
        "May": "Mai",
        "June": "Juin",
        "July": "Juillet",
        "August": "Août",
        "September": "Septembre",
        "October": "Octobre",
        "November": "Novembre",
        "December": "Décembre",
    }
    date_report = pd.to_datetime(date_report, format="%Y/%m/%d") + relativedelta(months=1)
    date_report = date_report.replace(day=11)
    ws_acceuil["E12"] = (
        "Période: "
        + dico_mois_fr[calendar.month_name[date_report.month]]
        + " "
        + str(date_report.year)
    )
    ws_acceuil["E15"] = "Extraction du 11/" + date_report.strftime("%m/%Y")

    if dest_file:
        wb_feedback_report.save(dest_file)

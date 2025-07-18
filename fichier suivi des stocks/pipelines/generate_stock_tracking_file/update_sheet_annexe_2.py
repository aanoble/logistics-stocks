import numpy as np
import pandas as pd
from efc.interfaces.iopenpyxl import OpenpyxlInterface
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter

from .constants import (
    ALIGNMENT,
    BODY_FONT,
    CENTER_ALIGNMENT,
    CS_FONT,
    DATE_STYLE,
    DICO_FORMULES_ANNEXE_2,
    DICO_RULES_ANNEXE_2,
    HEADER_FONT,
    THIN_BORDER,
)
from .utils import get_current_variable


def update_sheet_annexe_2(
    wb_temp: Workbook,
    df_plan_approv,
    date_report: str,
):
    """
    Met à jour la feuille 'Annexe 2 - Suivi des Stocks' dans le workbook donné en utilisant les données fournies.
    Args:
        wb_temp (Workbook): Le modèle de workbook à mettre à jour.
        df_plan_approv (DataFrame): Le dataframe contenant les informations de plan d'approvisionnement.
        date_report (str): La date de conception du rapport.
    Returns:
        Workbook: Le workbook mis à jour.
    """
    date_format = get_current_variable(date_report)[0]
    # Construction du data frame stock detaillé
    data = wb_temp["Stock detaille"].values
    cols = next(data)
    data = list(data)

    df_stock_detaille = pd.DataFrame(data, columns=cols)
    df_stock_detaille = df_stock_detaille.loc[df_stock_detaille["Code produit"].notna()]
    for col in ("Code produit", "Qté \nPhysique"):
        try:
            df_stock_detaille[col] = df_stock_detaille[col].astype(np.int64)
        except Exception:
            df_stock_detaille[col] = pd.to_numeric(df_stock_detaille[col], errors="coerce").astype(
                np.int64
            )

    del data, cols

    # Réceptions
    data = wb_temp["Receptions"].values
    cols = next(data)
    data = list(data)

    df_receptions = pd.DataFrame(data, columns=cols)
    df_receptions["Nouveau code"] = df_receptions["Nouveau code"].replace("ND", np.nan)
    try:
        df_receptions["Nouveau code"] = df_receptions["Nouveau code"].astype(float)
    except Exception:
        df_receptions["Nouveau code"] = pd.to_numeric(
            df_receptions["Nouveau code"], errors="coerce"
        ).astype(float)

    del data, cols

    ws_annexe_2 = wb_temp["Annexe 2 - Suivi des Stocks"]

    ws_annexe_2["AK1"].value = date_report
    ws_annexe_2["AK1"].font = CS_FONT
    ws_annexe_2["AP1"].value = "En cours"
    ws_annexe_2["AP1"].font = CS_FONT

    # Ajout de commentaire sur les colonnes pour lesquelles nous n'avons pas pu appliquer des formules

    ws_annexe_2["W2"].comment = Comment(
        """=MIN.SI.ENS('Stock detaille'!D:D; 'Stock detaille'!G:G;">0"; 'Stock detaille'!A:A; A5)""",
        author="author=",
    )  # Date de Péremption la plus proche (BRUTE)

    ws_annexe_2["AK3"].comment = Comment(
        """=SIERREUR(MIN.SI.ENS('Plan d''appro'!K:K; 'Plan d''appro'!A:A; A5; 'Plan d''appro'!K:K; ">"&$AK$1); "")""",
        author="author=",
    )  # Date Probable de Livraison

    ws_annexe_2["AL3"].comment = Comment(
        """=MIN.SI.ENS('Receptions'!F:F; 'Receptions'!C:C; A5; Receptions!J:J; "ok")""",
        author="author=",
    )  # Date effective de livraison

    special_font_bold_indices = [3] + list(range(9, 23)) + list(range(25, 28)) + list(range(29, 32))
    special_alignment_center_indices = [1] + list(range(6, 40))

    header_row = list(ws_annexe_2.iter_rows(min_row=3, max_row=3))[0]

    msd_number_col_indices = [
        cell.col_idx
        for cell in header_row
        if hasattr(cell, "col_idx")
        and cell.value is not None
        and ("MSD" in cell.value or "mois" in cell.value)
    ] + [26, 27]

    other_number_format = [25] + [
        cell.col_idx
        for cell in header_row
        if hasattr(cell, "col_idx")
        and cell.value is not None
        and (
            "SDU" in cell.value
            or "DMM" in cell.value
            or "CMM" in cell.value
            or "CONSO" in cell.value
            or "Qtité" in cell.value
        )
    ]

    interface = OpenpyxlInterface(wb=wb_temp, use_cache=True)
    interface.clear_cache()

    def format_cell_annexe_2(cell, col_idx):
        cell.border = THIN_BORDER
        cell.font = HEADER_FONT if col_idx in special_font_bold_indices else BODY_FONT
        cell.alignment = (
            CENTER_ALIGNMENT if col_idx in special_alignment_center_indices else ALIGNMENT
        )
        # Formattage des nombres
        if col_idx in msd_number_col_indices:
            cell.number_format = "0.0"
        if col_idx in other_number_format:
            cell.number_format = "#,##0"
        if col_idx in (24, 23, 38, 40):
            cell.style = DATE_STYLE

    max_row_annexe_1 = wb_temp["Annexe 1 - Consolidation"].max_row

    for start in range(5, max_row_annexe_1 + 1):
        for col_idx, formula in DICO_FORMULES_ANNEXE_2.items():
            if formula is None:
                cell = ws_annexe_2.cell(row=start, column=col_idx)
                cell.border = THIN_BORDER

            if col_idx not in (23, 38, 40):
                cell = ws_annexe_2.cell(row=start, column=col_idx, value=formula.format(start))
                format_cell_annexe_2(cell, col_idx)

            # Date de Péremption la plus proche (BRUTE)
            if col_idx == 23:
                # print(interface.calc_cell(f'C{start}', ws_annexe_2.title))
                value = df_stock_detaille.loc[
                    (df_stock_detaille["Qté \nPhysique"] > 0)
                    & (
                        df_stock_detaille["Code produit"]
                        == interface.calc_cell(f"A{start}", ws_annexe_2.title)
                    ),
                    "Date limite de consommation",
                ].min()
                cell = ws_annexe_2.cell(row=start, column=col_idx, value=value)
                # cell.style = DATE_STYLE
                format_cell_annexe_2(cell, col_idx)

            # Date Probable de Livraison
            eomonth = (
                pd.to_datetime(date_format).replace(day=1) + pd.offsets.MonthEnd(0)
            ).strftime("%Y-%m-%d")
            if col_idx == 37:
                value = (
                    df_plan_approv.loc[
                        (
                            df_plan_approv["Standard product code"]
                            == interface.calc_cell(f"A{start}", ws_annexe_2.title)
                        )
                        & (df_plan_approv["DATE"] > eomonth),
                        "DATE",
                    ].min()
                    if not df_plan_approv.loc[
                        (
                            df_plan_approv["Standard product code"]
                            == interface.calc_cell(f"A{start}", ws_annexe_2.title)
                        )
                        & (df_plan_approv["DATE"] > eomonth)
                    ].empty
                    else np.nan
                )

                cell = ws_annexe_2.cell(row=start, column=col_idx, value=value)
                # cell.style = DATE_STYLE
                format_cell_annexe_2(cell, col_idx)
                cell.number_format = "DD MMM YYYY"

            # Date effective de livraison
            if col_idx == 38:
                value = df_receptions.loc[
                    (
                        df_receptions["Nouveau code"]
                        == interface.calc_cell(f"A{start}", ws_annexe_2.title)
                    )
                    & (
                        (df_receptions["Date d'entrée en machine"] > date_format)
                        | df_receptions["Date d'entrée en machine"].isna()
                    ),
                    "Date de réception effective",
                ].max()
                cell = ws_annexe_2.cell(row=start, column=col_idx, value=value)
                # cell.style = DATE_STYLE
                format_cell_annexe_2(cell, col_idx)
                cell.number_format = "DD MMM YYYY"

        # Gestion du formattage des colonnes AN -> AR
        for col_idx in range(40, 45):
            cell = ws_annexe_2.cell(row=start, column=col_idx)
            cell.border = THIN_BORDER

    # header_row = list(ws_annexe_2.iter_rows(min_row=3, max_row=3))[0]
    statut_col_indices = [
        cell.col_idx
        for cell in header_row
        if hasattr(cell, "col_idx") and cell.value is not None and "STATUT" in cell.value
    ]
    # statut_col_indices
    f_rule_etat_stock = 'NOT(ISERROR(SEARCH("{etat_stock}", {col_letter_etat_stock}{index_start})))'

    for col_idx in statut_col_indices:
        col_letter_etat_stock = get_column_letter(col_idx)
        for etat_stock in DICO_RULES_ANNEXE_2:
            rule = Rule(
                type="containsText",
                operator="containsText",
                dxf=DICO_RULES_ANNEXE_2[etat_stock],
                text=etat_stock,
                formula=[
                    f_rule_etat_stock.format(
                        etat_stock=etat_stock,
                        col_letter_etat_stock=col_letter_etat_stock,
                        index_start=5,
                    )
                ],
            )
            ws_annexe_2.conditional_formatting.add(
                f"{col_letter_etat_stock}5:{col_letter_etat_stock}{start}", rule
            )

    return wb_temp

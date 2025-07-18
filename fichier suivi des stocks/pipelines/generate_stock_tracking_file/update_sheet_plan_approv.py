import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows

from .constants import CENTER_ALIGNMENT, DATE_STYLE, LEFT_ALIGNMENT


def update_sheet_plan_approv(wb_temp: Workbook, df_plan_approv: pd.DataFrame) -> Workbook:
    """Met à jour la feuille `Plan d'approvisionnement` en utilisant les données extraites de QAT.

    Args:
        wb_temp (Workbook): Workbook template.
        df_plan_approv (pd.DataFrame): DataFrame contenant les données prétraitées.

    Returns:
        Workbook: Le Workbook mis à jour.
    """

    df_plan_approv.columns = df_plan_approv.columns.str.strip()
    ws_plan_approv = wb_temp["Plan d'appro"]

    # Récupération des colonnes de l'en-tête et de leurs indices
    header_row = list(ws_plan_approv.iter_rows(max_col=column_index_from_string("R")))[0]
    cols_df_plan_approv = df_plan_approv.columns.tolist()

    dico_cols = {
        cell.value.strip(): (
            cell.column_letter,
            cell.col_idx,
            cols_df_plan_approv.index(cell.value.strip()),
        )
        for cell in header_row
        if cell.value and cell.value.strip() in cols_df_plan_approv
    }

    # Définition des styles
    font = Font(name="Arial Narrow", size=11)
    font_bold = Font(name="Arial Narrow", size=11, bold=True)
    fill = PatternFill(start_color="FFD9E1F2", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFFFF", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000")
    )

    number_format_cols = {
        "Standard product code",
        "Quantite",
        "Cout des Produits",
        "Couts du fret",
        "Couts totaux",
    }

    # Formules dynamiques
    dico_formules = {
        "J": "=I{row}*H{row}",  # Quantité harmonisée
        "P": '=TEXT(K{row},"mmm") & "-" &  YEAR(K{row})',  # Date mise à jour
        "Q": '=IF(G{row}="","",IF(G{row}="Reçu", 1, 0))',  # Reçu ?
        "S": "=I{row}*R{row}",  # Coût unitaire harmonisé
        "T": '=A{row}&"_"&K{row}',  # Concaténation du code et de la date
    }

    # Remplissage des cellules avec les données
    for row_idx, row in enumerate(
        dataframe_to_rows(df_plan_approv, index=False, header=False), start=2
    ):
        for col_name, (col_letter, col_idx, df_idx) in dico_cols.items():
            cell = ws_plan_approv.cell(row=row_idx, column=col_idx, value=row[df_idx])
            cell.font = font_bold if col_letter in {"A", "I", "J"} else font
            cell.border = border
            cell.alignment = (
                CENTER_ALIGNMENT
                if col_letter not in {"C", "E", "F", "G", "O", "P"}
                else LEFT_ALIGNMENT
            )
            cell.fill = fill if col_letter == "I" else fill_white

            if col_name == "DATE":
                cell.style = DATE_STYLE
                cell.font = font
                cell.border = border
                cell.fill = fill_white
            if col_name in number_format_cols:
                cell.number_format = "0"

        # Ajout des formules
        for col_letter, formula in dico_formules.items():
            cell = ws_plan_approv.cell(
                row=row_idx,
                column=column_index_from_string(col_letter),
                value=formula.format(row=row_idx),
            )
            cell.font = font_bold if col_letter in {"A", "I", "J"} else font
            cell.alignment = (
                LEFT_ALIGNMENT if col_letter not in {"J", "Q", "S"} else CENTER_ALIGNMENT
            )
            cell.border = border
            cell.fill = fill_white
            if col_letter == "J":
                cell.number_format = "0"
    return wb_temp

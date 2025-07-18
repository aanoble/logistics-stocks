import pandas as pd
from efc.interfaces.iopenpyxl import OpenpyxlInterface
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .constants import THIN_BORDER
from .utils import find_best_match, has_formula


def update_sheet_etat_stock(wb_temp, wb_fbr, programme):
    """
    Met à jour la feuille `Etat de Stock Periph` dans le classeur temporaire `wb_temp`
    en utilisant les données de la feuille `ETAT DU STOCK` du classeur `wb_fbr`.
    Args:
        wb_temp (openpyxl.Workbook): Le classeur temporaire où la feuille `Etat de stock Periph` sera mise à jour.
        wb_fbr (openpyxl.Workbook): Le classeur source contenant la feuille `ETAT DU STOCK` avec les données à copier.
        programme (str): Le programme spécifique pour lequel les données doivent être mises à jour.
    Returns:
        openpyxl.Workbook: Le classeur temporaire `wb_temp` mis à jour avec les nouvelles données.

    """
    ws_fbr = wb_fbr["ETAT DU STOCK"]
    ws_temp = wb_temp["Etat de stock Periph"]
    dico_cols = {}

    for row_base, row_temp in zip(
        ws_fbr.iter_rows(min_row=1, max_row=1), ws_temp.iter_rows(min_row=1, max_row=1)
    ):
        values = [cell.value for cell in row_base if cell.value is not None]
        col_idx_programme = values.index("PROGRAMME")
        for cell in row_temp:
            match_index = find_best_match(cell.value, values)
            if match_index is not None:
                dico_cols[match_index] = cell.col_idx

    interface = OpenpyxlInterface(wb=wb_fbr, use_cache=True)
    interface.clear_cache()
    index_row = 2
    for row in ws_fbr.iter_rows(min_row=2):
        if row[col_idx_programme].value != programme:
            continue
        for cell in row:
            if dico_cols.get(cell.col_idx) is None:
                continue
            elif has_formula(cell):
                result = interface.calc_cell(cell.coordinate, ws_fbr.title)
                new_cell = ws_temp.cell(row=index_row, column=dico_cols[cell.col_idx], value=result)
            else:
                new_cell = ws_temp.cell(
                    row=index_row, column=dico_cols[cell.col_idx], value=cell.value
                )

            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.alignment = cell.alignment.copy()

        ws_temp.row_dimensions[index_row].height = 27
        index_row += 1

    return wb_temp


def update_sheet_etat_stock(wb_temp, df_etat_stock: pd.DataFrame):
    """
    Met à jour la feuille "Etat de stock Periph" d'un classeur Excel avec les données d'un DataFrame.
    Cette fonction formate les données de la feuille "ETAT DU STOCK" du template en appliquant des styles,
    des alignements, des formats de nombre et des règles de mise en forme conditionnelle basées sur les valeurs
    du DataFrame fourni.
    Args:
        wb_temp (Workbook): Le classeur Excel à mettre à jour.
        df_etat_stock (pd.DataFrame): Le DataFrame contenant les données de l'état du stock.
    Returns:
        Workbook: Le classeur Excel mis à jour.
    """

    # Formule générale qui sera formatée pour les différentes règles
    f_rule_etat_stock = 'NOT(ISERROR(SEARCH("{etat_stock}", {col_letter_etat_stock}{index_start})))'

    ws_temp = wb_temp["Etat de stock Periph"]

    header_row = list(ws_temp.iter_rows(max_row=1))[0]

    df_etat_stock = df_etat_stock.rename(columns=lambda x: x.rstrip()).rename(
        columns={"BESOIN COMMANDE URGENTE": "BESOIN CMMMANDE URGENTE"}
    )

    if df_etat_stock.empty:
        # Pour les programmes tels que le PNLT qui font des rapportages chaque trimestre
        return wb_temp

    cols_df_etat_stock = list(df_etat_stock.columns)

    dico_cols = {}

    for cell in header_row:
        # _ = cell.value.replace('  ', ' ').rstrip()
        if cell.value.rstrip() in cols_df_etat_stock:
            dico_cols[cell.value.rstrip()] = [
                cell.column_letter,
                cell.col_idx,
                cols_df_etat_stock.index(cell.value.rstrip()),
            ]

    font_one = Font(name="Arial Narrow", size=11, bold=True)
    font_two = Font(name="Arial Narrow", size=11)
    fill = PatternFill(start_color="FFDDEBF7", fill_type="solid")
    alignment_one = Alignment(horizontal=None, vertical="center", wrap_text=True)
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    special_font_cols = {
        "CODE",
        "CODE  ETS",
        "CMM gestionnaire",
        "QUANTITE PROPOSEE",
        "MSD",
        "ETAT DU STOCK",
        "BESOIN CMMMANDE URGENTE",
        "BESOIN TRANSFERT IN",
        "QUANTITE A TRANSFERER OUT",
    }

    special_alignment_cols = {"Code_Pro", "CODE", "PROGRAMME", "CODE  ETS", "CATEGORIE PRODUIT"}
    special_alignment_indices = range(13, 29)

    number_format_cols = {
        "CMM gestionnaire": "#,##0",
        "Code_Pro": "0.00",
        "CATEGORIE PRODUIT": "0.00",
        "MSD": "0.00",
    }
    fill_cols = {
        "Code_Pro",
        "PROGRAMME",
        "CATEGORIE PRODUIT",
        "CMM gestionnaire",
        "MSD",
        "ETAT DU STOCK",
        "BESOIN CMMMANDE URGENTE",
        "BESOIN TRANSFERT IN",
        "QUANTITE A TRANSFERER OUT",
    }

    for start, row in enumerate(
        dataframe_to_rows(df_etat_stock, index=False, header=False), start=2
    ):
        ws_temp.row_dimensions[start].height = 25.2
        for col, element in dico_cols.items():
            cell = ws_temp.cell(row=start, column=element[1], value=row[element[2]])
            cell.border = THIN_BORDER
            cell.font = font_one if col in special_font_cols else font_two
            cell.alignment = (
                alignment_center
                if col in special_alignment_cols or element[1] in special_alignment_indices
                else alignment_one
            )
            if col in number_format_cols:
                cell.number_format = number_format_cols[col]
            if col in fill_cols:
                cell.fill = fill

    # Add rule in column Categorie du produit
    rule_categorie_produit = CellIsRule(
        operator="equal",
        formula=['"Produit traceur"'],
        fill=PatternFill(start_color="FF8EA9DB", end_color="FF8EA9DB", fill_type="solid"),
        font=Font(name="Arial Narrow", size=11, bold=True),
    )

    ws_temp.conditional_formatting.add(
        f"{dico_cols['CATEGORIE PRODUIT'][0]}3:{dico_cols['CATEGORIE PRODUIT'][0]}{start}",
        rule_categorie_produit,
    )

    # Add rule in column Etat du stock
    col_letter_etat_stock = dico_cols["ETAT DU STOCK"][0]

    dico_dxf = {
        "STOCK DORMANT": DifferentialStyle(
            fill=PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True, color="FFFFFFFF"),
        ),
        "SURSTOCK": DifferentialStyle(
            fill=PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True),
        ),
        "RUPTURE": DifferentialStyle(
            fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True, color="FFFFFFFF"),
        ),
        "BIEN STOCKE": DifferentialStyle(
            fill=PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True),
        ),
        "EN BAS DU PCU": DifferentialStyle(
            fill=PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True),
        ),
        "ENTRE PCU et MIN": DifferentialStyle(
            fill=PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True),
        ),
    }
    for etat_stock in dico_dxf:
        rule = Rule(
            type="containsText",
            operator="containsText",
            dxf=dico_dxf[etat_stock],
            text=etat_stock,
            formula=[
                f_rule_etat_stock.format(
                    etat_stock=etat_stock,
                    col_letter_etat_stock=col_letter_etat_stock,
                    index_start=3,
                )
            ],
        )

        ws_temp.conditional_formatting.add(
            f"{col_letter_etat_stock}3:{col_letter_etat_stock}{start}", rule
        )

    ws_temp.sheet_state = "hidden"

    return wb_temp


def update_sheet_stock_region(wb_temp, wb_fbr, programme):
    """
    Met à jour la feuille `StockParRegion` dans un classeur Excel temporaire en utilisant les données d'un autre classeur Excel.
    Args:
        wb_temp (openpyxl.Workbook): Le classeur Excel temporaire dans lequel la feuille sera mise à jour.
        wb_fbr (openpyxl.Workbook): Le classeur Excel contenant les données de la feuille `StockParRegion`.
        programme (str): Le nom du programme pour lequel les données doivent être mises à jour.
    Returns:
        openpyxl.Workbook: Le classeur Excel temporaire mis à jour.
    """
    ws_fbr = wb_fbr["StockParRegion"]
    ws_temp_old = wb_temp["StockParRegion"]
    ws_temp = wb_temp.create_sheet(
        "New StockParRegion", wb_temp.sheetnames.index("StockParRegion") + 1
    )

    row_base = list(ws_fbr.iter_rows(min_row=2, max_row=2))[0]
    values = [cell.value for cell in row_base if cell.value is not None]

    col_idx_programme = values.index("Programme")

    interface = OpenpyxlInterface(wb=wb_fbr, use_cache=True)
    interface.clear_cache()
    index_row = 1
    alignment_code = Alignment(horizontal="center", vertical="center")

    for start, row in enumerate(ws_fbr.iter_rows(), start=1):
        if start > 3 and row[col_idx_programme].value != programme:
            continue
        for cell in row:
            try:
                if has_formula(cell):
                    result = interface.calc_cell(cell.coordinate, ws_fbr.title)
                    new_cell = ws_temp.cell(row=index_row, column=cell.col_idx, value=result)
                else:
                    new_cell = (
                        ws_temp.cell(row=index_row, column=cell.col_idx, value=cell.value)
                        if cell.col_idx != 1
                        else ws_temp.cell(
                            row=index_row,
                            column=cell.col_idx,
                            value=int(str(cell.value).replace(f"_{programme}", "")),
                        )
                    )
            except AttributeError:
                continue

            except ValueError:
                new_cell = ws_temp.cell(row=index_row, column=cell.col_idx, value=cell.value)

            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.alignment = cell.alignment.copy() if cell.col_idx != 1 else alignment_code

        ws_temp.row_dimensions[index_row].height = 18
        index_row += 1

    ws_temp.delete_rows(1)

    for merged_cell in ws_fbr.merged_cells.ranges:
        ws_temp.merge_cells(str(merged_cell).replace("2", "1").replace("3", "2"))

    for row in ws_temp_old.iter_rows(max_row=1):
        for cell in row:
            coordinate = cell.coordinate.strip("1")
            ws_temp.column_dimensions[coordinate].width = ws_temp_old.column_dimensions[
                coordinate
            ].width

    wb_temp.remove(ws_temp_old)

    ws_temp.title = "StockParRegion"

    # Gestion de la mise en forme conditionnelle des status
    header_row = list(ws_temp.iter_rows(min_row=2, max_row=2))[0]
    statut_col_indices = [
        cell.col_idx
        for cell in header_row
        if hasattr(cell, "col_idx") and cell.value is not None and "STATUT" in cell.value
    ]

    dico_dxf = {
        "STOCK DORMANT": DifferentialStyle(
            fill=PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid"),
            font=Font(name="Arial Narrow", size=11, bold=True, color="FFFFFFFF"),
        ),
        "SURSTOCK": DifferentialStyle(
            fill=PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid"),
            font=Font(name="Arial Narrow", size=10, bold=True),
        ),
        "RUPTURE": DifferentialStyle(
            fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
            font=Font(name="Arial Narrow", size=10, bold=True, color="FFFFFFFF"),
        ),
        "BIEN STOCKE": DifferentialStyle(
            fill=PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid"),
            font=Font(name="Arial Narrow", size=10, bold=True),
        ),
        "SOUS-STOCK": DifferentialStyle(
            fill=PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),
            font=Font(name="Arial Narrow", size=10, bold=True),
        ),
        "NA": DifferentialStyle(
            fill=PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"),
            font=Font(name="Arial Narrow", size=10, bold=True),
        ),
    }

    f_rule_etat_stock = 'NOT(ISERROR(SEARCH("{etat_stock}", {col_letter_etat_stock}{index_start})))'
    row_max = ws_temp.max_row
    for col_idx in statut_col_indices:
        col_letter_etat_stock = get_column_letter(col_idx)
        for etat_stock in dico_dxf:
            rule = Rule(
                type="containsText",
                operator="containsText",
                dxf=dico_dxf[etat_stock],
                text=etat_stock,
                formula=[
                    f_rule_etat_stock.format(
                        etat_stock=etat_stock,
                        col_letter_etat_stock=col_letter_etat_stock,
                        index_start=3,
                    )
                ],
            )
            try:  # Pour les programmes tels que le PNLT il y a des exceptions compte tenu du fait qu'ils font des rapportages chaque trimestre
                ws_temp.conditional_formatting.add(
                    f"{col_letter_etat_stock}3:{col_letter_etat_stock}{row_max}", rule
                )
            except Exception:
                continue

    for row in range(3, row_max + 1):
        ws_temp.row_dimensions[row].height = 34

    ws_temp.delete_cols(column_index_from_string("BY"))

    ws_temp.sheet_state = "hidden"

    return wb_temp

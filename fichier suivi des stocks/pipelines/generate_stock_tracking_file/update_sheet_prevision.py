import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows

from .constants import (
    ALIGNMENT,
    BODY_FONT,
    CENTER_ALIGNMENT,
    DICO_FORMULES_PREVISION,
    DICO_RULES_PREVISION,
    HEADER_FONT,
)
from .utils import get_current_variable


def update_sheet_prevision(
    wb_temp: Workbook,
    date_report: str,
    df_produit: pd.DataFrame,
):
    """Mise à jour de la feuille 'Prévision' dans le workbook donné en utilisant les données fournies.
    Args:
        wb_temp (Workbook): objet workbook template
        date_report (str): date de rapportage
        df_produit (pd.DataFrame): dataframe comportant la liste des produits
    """

    date_format = get_current_variable(date_report)[0]

    ws_prevision = wb_temp["Prévision"]
    ws_prevision["J6"].value = date_format
    fill = PatternFill(start_color="FFF2F2F2", fill_type="solid")
    medium_border = Border(left=Side(style="medium", color="000000"))
    BODY_FONT.size = 11
    HEADER_FONT.size = 11

    def format_cell_prevision(cell, center=True, bold=True):
        cell.font = HEADER_FONT if bold else BODY_FONT
        cell.alignment = CENTER_ALIGNMENT if center else ALIGNMENT
        cell.fill = fill

    for start, row in enumerate(dataframe_to_rows(df_produit, index=False, header=False), start=8):
        col_values = {
            "E": (row[0], "0"),  # Code produit
            "F": (row[1], None),  # Type ou catégorie
            "G": (row[2], None),  # Désignation
        }

        for col, (value, num_format) in col_values.items():
            cell = ws_prevision.cell(row=start, column=column_index_from_string(col), value=value)
            format_cell_prevision(cell, bold=False)

            if num_format:
                cell.number_format = num_format

        for col, formula in DICO_FORMULES_PREVISION.items():
            cell = ws_prevision.cell(
                row=start, column=column_index_from_string(col.upper()), value=formula.format(start)
            )
            if col in {"x", "y", "z", "aq", "ar", "bg", "bh", "bw", "bx"}:
                format_cell_prevision(cell, bold=False)
            else:
                format_cell_prevision(cell, center=True)

            # Définition des formats numériques selon les plages de colonnes
            col_idx = cell.col_idx
            if col in {"x", "aq", "bg", "bw"}:
                cell.number_format = "0"
            elif 7 <= col_idx < 23 or 27 <= col_idx < 44:
                cell.number_format = "#,##0"
            elif 61 <= col_idx <= 73:
                cell.number_format = "0.0"
            elif 77 <= col_idx <= 89:
                cell.number_format = '"$"#,##0.00'

            # Gestion des alignements spécifiques
            if col in {"h", "i", "aa", "ab"}:
                cell.alignment = Alignment(horizontal="right", vertical="center")

        # Gestion des bordures et mises en forme des autres cellules
        for col in {"D", "W", "AP", "BF", "BV"}:
            cell = ws_prevision.cell(row=start, column=column_index_from_string(col))
            format_cell_prevision(cell)
            cell.border = medium_border
        cell = ws_prevision.cell(row=start, column=column_index_from_string("CL"))
        cell.border = medium_border
    # Application des règles de mise en forme conditionnelle
    for col in DICO_FORMULES_PREVISION.keys():
        col_idx = column_index_from_string(col.upper())
        col_range = f"{col}8:{col}{start}"

        if col in {"h", "i", "x", "y", "z", "aa", "ab", "aq", "ar", "bg", "bh", "bw", "bx"}:
            continue

        if col_idx < column_index_from_string("W"):
            rules = [
                "rule_equal_nd",
                "rule_equal_zero",
                "rule_less_than_third",
                "rule_between_third_and_eight",
                "rule_greater_than_eight",
            ]
        elif col_idx < column_index_from_string("AP"):
            rules = [
                "rule_equal_nd",
                "rule_equal_zero",
                "rule_less_than_five",
                "rule_between_five_and_twelve",
                "rule_greater_than_twelve",
            ]
        else:
            rules = ["rule_greater_than_zero", "rule_equal_empty"]

        for rule in rules:
            ws_prevision.conditional_formatting.add(col_range, DICO_RULES_PREVISION[rule])

    return wb_temp

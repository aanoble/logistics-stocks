import re
from pathlib import Path

import numpy as np
from openpyxl import Workbook
import pandas as pd
from compute_indicators.file_utils import process_etat_stock_npsp
from compute_indicators.utils import check_if_sheet_name_in_file
from efc.interfaces.iopenpyxl import OpenpyxlInterface  # type: ignore
from generate_stock_tracking_file.utils import has_formula


def get_data_from_sheet(
    fp_suivi_stock: Path,
    sheet_name: str,
    sheetnames: list[str],
    date_report: str | None = None,
    programme: str | None = None,
    src_wb: Workbook | None = None,
) -> pd.DataFrame:
    """Get data from a specific sheet in the stock tracking file.

    Args:
        fp_suivi_stock (Path): Path to the stock tracking file.
        sheet_name (str): Name of the sheet to extract data from.
        sheetnames (list[str]): List of sheet names in the workbook.
        date_report (str | None): Date of the report, if applicable.
        programme (str | None): Programme name, if applicable.
        src_wb: Source workbook, if needed for calculations.

    Returns:
        pd.DataFrame: DataFrame containing the data from the specified sheet.

    Raises:
        ValueError: If the sheet name is not recognized.
    """
    if sheet_name == "Etat de stock":
        sheet_stock_npsp = check_if_sheet_name_in_file("Etat de stock", sheetnames)
        assert sheet_stock_npsp is not None, print(
            f"La feuille `Etat de stock` n'est pas dans la liste {sheetnames} du classeur excel"
        )

        df_etat_stock_npsp = pd.read_excel(fp_suivi_stock, sheet_name=sheet_stock_npsp, skiprows=4)
        df_etat_stock_npsp = df_etat_stock_npsp.loc[df_etat_stock_npsp["Nouveau code"].notna()]
        df_etat_stock_npsp = process_etat_stock_npsp(df_etat_stock_npsp, date_report, programme)

        return df_etat_stock_npsp

    elif sheet_name == "Stock detaille":
        sheet_stock_detaille = check_if_sheet_name_in_file("Stock detaille", sheetnames)
        assert sheet_stock_detaille is not None, print(
            f"La feuille `Stock detaille` n'est pas dans la liste {sheetnames} du classeur excel"
        )

        df_stock_detaille = pd.read_excel(fp_suivi_stock, sheet_name=sheet_stock_detaille)

        max_date_year = pd.Timestamp.max.year

        try:
            df_stock_detaille["Date limite de consommation"] = df_stock_detaille[
                "Date limite de consommation"
            ].apply(lambda x: x if x.year < max_date_year else x.replace(year=max_date_year - 1))
        except Exception:
            df_stock_detaille["Date limite de consommation"] = df_stock_detaille[
                "Date limite de consommation"
            ].str.strip()

            df_stock_detaille["Date limite de consommation"] = df_stock_detaille[
                "Date limite de consommation"
            ].apply(
                lambda x: x
                if int(x[-4:]) < max_date_year
                else x.replace(x[-4:], str(max_date_year - 1))
            )

        df_stock_detaille["Date limite de consommation"] = pd.to_datetime(
            df_stock_detaille["Date limite de consommation"], format="%d/%m/%Y"
        )

        return df_stock_detaille

    elif sheet_name == "Distribution X3":
        sheet_distribution_x3 = check_if_sheet_name_in_file("Distribution X3", sheetnames)
        assert sheet_distribution_x3 is not None, print(
            f"La feuille `Distribution X3` n'est pas dans la liste {sheetnames} du classeur excel"
        )

        df_distribution = pd.read_excel(fp_suivi_stock, sheet_name=sheet_distribution_x3)

        return df_distribution

    elif sheet_name == "Receptions":
        sheet_reception = check_if_sheet_name_in_file("Receptions", sheetnames)
        assert sheet_reception is not None, print(
            f"La feuille `Receptions` n'est pas dans la liste {sheetnames} du classeur excel"
        )

        df_receptions = pd.read_excel(fp_suivi_stock, sheet_name=sheet_reception)

        return df_receptions

    elif sheet_name == "PPI":
        sheet_ppi = check_if_sheet_name_in_file("PPI", sheetnames)
        assert sheet_ppi is not None, print(
            f"La feuille `PPI` n'est pas dans la liste {sheetnames} du classeur excel"
        )

        df_ppi = pd.read_excel(fp_suivi_stock, sheet_name=sheet_ppi, skiprows=2)

        return df_ppi

    elif sheet_name == "Prelèvement CQ":
        sheet_prelev = check_if_sheet_name_in_file("Prelèvement CQ", sheetnames)
        assert sheet_prelev is not None, print(
            f"La feuille `Prelèvement CQ` n'est pas dans la liste {sheetnames} du classeur excel"
        )

        df_prelevement = pd.read_excel(fp_suivi_stock, sheet_name=sheet_prelev, skiprows=2)

        return df_prelevement

    elif sheet_name == "Plan d'appro":
        import locale

        locale.setlocale(locale.LC_TIME, "fr_FR.UTF-8")

        sheet_approv = check_if_sheet_name_in_file("Plan d'appro", sheetnames)
        assert sheet_approv is not None, print(
            f"La feuille `Plan d'appro` n'est pas dans la liste {sheetnames} du classeur excel"
        )

        interface = OpenpyxlInterface(wb=src_wb, use_cache=True)
        interface.clear_cache()
        data_list = []
        for row in src_wb[sheet_approv].iter_rows(min_row=0, max_col=19):
            data = []
            for cell in row:
                if cell.column_letter == "P":
                    continue
                if has_formula(cell):
                    result = interface.calc_cell(cell.coordinate, sheet_approv)
                    data.append(result)
                else:
                    data.append(cell.value)
            data_list.append(data)

        df_plan_approv = pd.DataFrame(data_list[1:], columns=data_list[0])

        # df_plan_approv = pd.read_excel(fp_suivi_stock, sheet_name=sheet_approv, engine="openpyxl")
        df_plan_approv.columns = df_plan_approv.columns.str.strip()
        df_plan_approv["Date updated"] = df_plan_approv["DATE"].dt.strftime("%B-%Y")

        return df_plan_approv

    elif sheet_name == "Statut Produits":
        sheet_statut_prod = check_if_sheet_name_in_file("Statut Produits", sheetnames)
        assert sheet_statut_prod is not None, print(
            f"La feuille `Statut Produits` n'est pas dans la liste {sheetnames} du classeur excel"
        )

        df_statut_prod = pd.read_excel(fp_suivi_stock, sheet_name=sheet_statut_prod, skiprows=1)
        df_statut_prod["programme"] = programme

        return df_statut_prod
    elif sheet_name == "Annexe 1 - Consolidation":
        sheet_annexe_1 = check_if_sheet_name_in_file("Annexe 1 - Consolidation", sheetnames)

        assert sheet_annexe_1 is not None, print(
            f"La feuille `Annexe 1 - Consolidation` n'est pas dans la liste {sheetnames} du classeur excel"
        )

        df_etat_stock = pd.read_excel(
            fp_suivi_stock, sheet_name=sheet_annexe_1, skiprows=2, usecols="A:T", engine="openpyxl"
        ).dropna(how="all")

        COLUMN_MAPPING = {
            "Stock Théorique fin": "stock_theorique_mois_precedent",
        }
        df_etat_stock.rename(
            columns={"CODE": "code_produit"},
            inplace=True,
        )

        df_etat_stock.rename(
            columns=lambda col: next(
                (v for k, v in COLUMN_MAPPING.items() if re.search(k, col, re.I)), col
            )
            if not col.endswith("SAGE") and not col.endswith("Final Attendu")
            else col,
            inplace=True,
        )

        interface = OpenpyxlInterface(wb=src_wb, use_cache=True)
        interface.clear_cache()
        data_list = []
        for row in src_wb[sheet_annexe_1].iter_rows(min_row=5, max_col=20):
            data = []
            for cell in row:
                if has_formula(cell):
                    result = interface.calc_cell(cell.coordinate, sheet_annexe_1)
                    data.append(result)
                else:
                    data.append(cell.value)
            data_list.append(data)

        if data_list:
            df_etat_stock = pd.DataFrame(data_list, columns=df_etat_stock.columns)

            df_etat_stock.fillna(np.nan, inplace=True)

        return df_etat_stock
    else:
        raise ValueError(f"Le nom de la feuille `{sheet_name}` n'est pas reconnu.")

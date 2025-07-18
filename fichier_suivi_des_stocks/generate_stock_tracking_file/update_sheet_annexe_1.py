from typing import Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import Engine

from .constants import (
    ALERT_FONT,
    ALIGNMENT,
    BODY_FONT,
    CENTER_ALIGNMENT,
    CS_FONT,
    DICO_RULES_ANNEXE_1,
    HEADER_FONT,
    LIGHT_BLUE_FILL,
    QUERY_ANNEXE_1_CMM_GLOBAL,
    QUERY_ANNEXE_1_CMM_HISTO,
    QUERY_ANNEXE_1_DMM_GLOBAL,
    QUERY_ANNEXE_1_DMM_HISTO,
    QUERY_ANNEXE_1_PRODUIT,
    THIN_BORDER,
)
from .utils import get_current_variable


def apply_format_on_dmm_cell(cell, cross=False):
    """
    Cette fonction permet d'appliquer la mise en forme des cellules de la Distribution mensuelles Enregistrées
    """
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGNMENT
    if not cross:
        cell.number_format = "#,##0"
        cell.font = BODY_FONT
    else:
        cell.font = HEADER_FONT
        cell.fill = LIGHT_BLUE_FILL


def update_first_informations(
    wb_temp: Workbook,
    df_produit: pd.DataFrame,
    date_report: str,
) -> Tuple[Worksheet, pd.DataFrame]:
    """
    Met à jour les premières informations de la feuille "Annexe 1 - Consolidation" dans le classeur Excel.
    Args:
        wb_temp (Workbook): Le classeur Excel temporaire.
        df_produit (pd.DataFrame): Le DataFrame contenant les informations des produits.
        date_report (str): La date du rapport sous forme de chaîne de caractères.
    Returns:
        Tuple[Worksheet, pd.DataFrame]: La feuille de calcul mise à jour et le DataFrame des produits modifié.
    """

    global date_format, month_year_str, prev_month_year_str

    date_format, month_year_str, prev_month_year_str = get_current_variable(date_report)

    ws_annexe_1 = wb_temp["Annexe 1 - Consolidation"]

    ws_annexe_1["I2"].value = date_format
    ws_annexe_1["I3"].value = "Stock Théorique fin " + prev_month_year_str

    header_row = list(ws_annexe_1.iter_rows(min_row=3, max_row=3, max_col=9))[0]

    df_produit = df_produit.rename(
        columns={
            "code_produit": "CODE",
            "ancien_code": "Ancien code",
            "categorie": "CATEGORIE",
            "designation": "DESIGNATION DU PRODUIT",
            "type_produit": "Type",
            "unit_niveau_central": "Unité niv Central",
            "unit_niveau_peripherique": "Unité niv Périphérique",
            "facteur_de_conversion": "Facteur de conversion \n(De la centrale à la périphérie)",
            "stock_theorique_mois_precedent": "Stock Théorique fin " + prev_month_year_str,
        }
    ).sort_values("CODE")

    cols_df_prod = list(df_produit.columns)
    dico_cols = {}

    for cell in header_row:
        if cell.value is None:
            continue
        if cell.value.rstrip() in cols_df_prod:
            dico_cols[cell.value.rstrip()] = [
                cell.column_letter,
                cell.col_idx,
                cols_df_prod.index(cell.value.rstrip()),
                max([len(val) for val in df_produit[cell.value.rstrip()].astype(str)]),
            ]

    for start, row in enumerate(dataframe_to_rows(df_produit, index=False, header=False), start=5):
        for col, element in dico_cols.items():
            cell = ws_annexe_1.cell(row=start, column=element[1], value=row[element[2]])
            # Gestion de la mise en forme
            cell.border = THIN_BORDER
            cell.font = (
                HEADER_FONT
                if (col == "CATEGORIE") or ("Stock Théorique" in col)
                else ALERT_FONT
                if col == "Facteur de conversion \n(De la centrale à la périphérie)"
                else BODY_FONT
            )
            cell.alignment = (
                CENTER_ALIGNMENT
                if col
                in (
                    "CATEGORIE",
                    "CODE",
                    "Facteur de conversion \n(De la centrale à la périphérie)",
                    "Unité niv Central",
                )
                or ("Stock Théorique" in col)
                else ALIGNMENT
            )
            if "Stock Théorique" in col:
                cell.number_format = "#,##0"

    ws_annexe_1.conditional_formatting.add(
        f"{dico_cols['Type'][0]}5:{dico_cols['Type'][0]}{start}",
        DICO_RULES_ANNEXE_1["rule_type_produit"],
    )

    header_row = list(ws_annexe_1.iter_rows(min_row=3, max_row=3, min_col=10, max_col=20))[0]

    dico_cols = {cell.value: cell.col_idx for cell in header_row}

    dico_formules = {
        "Distribution effectuée": "=SUMIFS('Distribution X3'!M:M,'Distribution X3'!I:I,A{0})",
        "Quantité reçue entrée en stock": '=SUMIFS(Receptions!H:H, Receptions!C:C,A{0}, Receptions!J:J,"ok")',
        "Quantité de PPI": "=SUMIFS(PPI!F:F,PPI!A:A,A{0})",
        "Quantité prélévée en Contrôle Qualité (CQ)": "=SUMIFS('Prelèvement CQ'!F:F,'Prelèvement CQ'!A:A,A{0})",
        "Stock Théorique Final SAGE": "=SUMIFS('Stock detaille'!G:G,'Stock detaille'!A:A,A{0})",
        "Stock Théorique Final Attendu": '=IF(O{0}="","",I{0}-J{0}+K{0}-L{0}-M{0}+N{0})',
        "ECARTS": "=O{0}-P{0}",
    }

    for start, i in enumerate(range(df_produit.shape[0]), start=5):
        for col, col_idx in dico_cols.items():
            if col not in dico_formules:
                cell = ws_annexe_1.cell(row=start, column=col_idx)
                cell.border = THIN_BORDER
            else:
                cell = ws_annexe_1.cell(
                    row=start, column=col_idx, value=dico_formules.get(col).format(start)
                )
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGNMENT
                cell.font = (
                    HEADER_FONT if col in ("Stock Théorique Final SAGE", "ECARTS") else BODY_FONT
                )
                cell.number_format = "#,##0"

    col_letter_ecart = get_column_letter(dico_cols["ECARTS"])

    ws_annexe_1.conditional_formatting.add(
        f"{col_letter_ecart}4:{col_letter_ecart}{start}", DICO_RULES_ANNEXE_1["rule_less_than_zero"]
    )
    ws_annexe_1.conditional_formatting.add(
        f"{col_letter_ecart}4:{col_letter_ecart}{start}",
        DICO_RULES_ANNEXE_1["rule_greater_than_zero"],
    )
    ws_annexe_1.conditional_formatting.add(
        f"{col_letter_ecart}4:{col_letter_ecart}{start}", DICO_RULES_ANNEXE_1["rule_equal_zero"]
    )
    for row in range(5, start + 1):
        ws_annexe_1.row_dimensions[row].height = 43

    return ws_annexe_1, df_produit


def update_dmm_informations_on_sheet(
    wb_temp,
    df_dmm_global,
    df_dmm_histo,
    df_produit,
    date_report,
) -> Tuple[Worksheet, pd.DataFrame]:
    """
    Met à jour les informations DMM sur la feuille de calcul "Annexe 1 - Consolidation".
    Args:
        wb_temp (Workbook): Le classeur Excel temporaire.
        date_report (datetime): La date du rapport.
        df_dmm_global (DataFrame): Le DataFrame contenant les informations globales de Distributions.
        df_dmm_histo (DataFrame): Le DataFrame contenant les informations historiques des Distributions sélectionnées.
        df_produit (DataFrame): Le DataFrame contenant les informations des produits.
    Returns:
        Tuple[Worksheet, pd.DataFrame]: La feuille de calcul mise à jour et le DataFrame des produits.
    """
    ws_annexe_1, df_produit = update_first_informations(
        wb_temp,
        df_produit,
        date_report,
    )

    ws_annexe_1 = wb_temp["Annexe 1 - Consolidation"]
    ws_annexe_1["BG2"].value = "Détermination de la DMM au mois de " + month_year_str

    # Modification des plages historique des valeurs des DMM et CMM
    dates_list = pd.date_range(
        start=f"{date_format.year - 1}-07-01", end=f"{date_format.year}-12-01", freq="MS"
    ).tolist()
    for row in ws_annexe_1.iter_rows(min_row=4, max_row=4, min_col=22, max_col=57):
        for cell in row:
            try:
                value = dates_list[0]
                if value:
                    ws_annexe_1.cell(row=cell.row, column=cell.col_idx, value=value)
                    ws_annexe_1.cell(row=cell.row, column=cell.col_idx + 42, value=value)
                dates_list.pop(0)
            except Exception:
                continue

    df_dmm_to_sheet = pd.pivot_table(
        df_dmm_global,
        values="dmm",
        index=["id_dim_produit_stock_track_pk"],
        columns="date_report",
    ).reset_index()

    df_dmm_to_sheet = df_dmm_to_sheet.rename(
        columns=lambda x: str(x).replace(" 00:00:00", "").lstrip() if " 00:00:00" in str(x) else x
    )

    df_dmm_to_sheet.columns.name = ""

    assert (
        df_produit.merge(df_dmm_to_sheet, on="id_dim_produit_stock_track_pk", how="left").shape[0]
        == df_produit.shape[0]
    )

    # Cette jointure pour garder l'ordre des données
    df_dmm_to_sheet = df_produit.merge(
        df_dmm_to_sheet, on="id_dim_produit_stock_track_pk", how="left"
    ).sort_values("CODE")[df_dmm_to_sheet.columns.tolist()]

    cols_df_dmm = df_dmm_to_sheet.columns.tolist()

    header_row = list(ws_annexe_1.iter_rows(min_row=4, max_row=4, min_col=22, max_col=57))[0]
    dico_cols = {}

    for cell in header_row:
        if cell.value is None or not hasattr(cell, "col_idx"):
            continue
        elif cell.value.strftime("%Y-%m-%d") in cols_df_dmm:
            dico_cols[cell.value.strftime("%Y-%m-%d")] = [
                cell.column_letter,
                cell.col_idx,
                cols_df_dmm.index(cell.value.strftime("%Y-%m-%d")),
                max(
                    [
                        len(val)
                        for val in df_dmm_to_sheet[cell.value.strftime("%Y-%m-%d")].astype(str)
                    ]
                ),
            ]
        else:
            dico_cols[cell.value.strftime("%Y-%m-%d")] = cell.col_idx

    # Gestion des autres colonnes renseignant sur la DMM du mois courant
    header_row = list(ws_annexe_1.iter_rows(min_row=3, max_row=3, min_col=59, max_col=62))[0]

    dico_cols.update({cell.value: cell.col_idx for cell in header_row})

    dico_formules = {
        "Nbre de mois de considérés": '=COUNTIFS(V{0}:BE{0},"X")',
        "Distributions enregistrées sur les mois de considérés": '=SUM(IF(W{0}="X",V{0},0),IF(Y{0}="X",X{0},0),IF(AA{0}="X",Z{0},0),IF(AC{0}="X",AB{0},0),IF(AE{0}="X",AD{0},0),IF(AG{0}="X",AF{0},0),IF(AI{0}="X",AH{0},0),IF(AK{0}="X",AJ{0},0),IF(AM{0}="X",AL{0},0),IF(AO{0}="X",AN{0},0),IF(AQ{0}="X",AP{0},0),IF(AS{0}="X",AR{0},0),IF(AU{0}="X",AT{0},0),IF(AW{0}="X",AV{0},0),IF(AY{0}="X",AX{0},0),IF(BA{0}="X",AZ{0},0),IF(BC{0}="X",BB{0},0), IF(BE{0}="X",BD{0},0))',
        "DMM Calculée \n(à valider pour ce mois)": "=IFERROR(BH{0}/BG{0},0)",
    }

    # print(dico_cols)
    for start, row in enumerate(
        dataframe_to_rows(df_dmm_to_sheet, index=False, header=False), start=5
    ):
        for col, element in dico_cols.items():
            if col == date_format.strftime("%Y-%m-%d"):
                column_element = element if not isinstance(element, list) else element[1]
                cell = ws_annexe_1.cell(row=start, column=column_element, value=f"=$J{start}")
                apply_format_on_dmm_cell(cell)

                # Il faut également faire des mise à jour de la cellule droite
                cell = ws_annexe_1.cell(row=start, column=column_element + 1, value="X")
                apply_format_on_dmm_cell(cell, True)
            elif not isinstance(element, int):
                cell = ws_annexe_1.cell(row=start, column=element[1], value=row[element[2]])
                apply_format_on_dmm_cell(cell)
                # Gestion de la Céllule de droite !!!!
                # Ici il faut avoir le nombre de mois considérés sur les données courant
                # Recherche de la valeur pour le coching de la cellule suivante
                mois_considere = df_dmm_histo.loc[
                    (df_dmm_histo.id_dim_produit_stock_track_pk == row[0]), "date_report_prev"
                ].values

                mois_considere = [date.strftime("%Y-%m-%d") for date in mois_considere]

                # print(col ,mois_considere, col in mois_considere)

                # print(mois_considere, pd.date_range(end=date_format, periods=int(mois_considere), freq='MS'))
                val = "X" if col in mois_considere else None
                cell = ws_annexe_1.cell(row=start, column=element[1] + 1, value=val)
                apply_format_on_dmm_cell(cell, True)

            elif col in dico_formules:
                cell = ws_annexe_1.cell(
                    row=start, column=element, value=dico_formules.get(col).format(start)
                )
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGNMENT
                cell.font = (
                    BODY_FONT if col != "DMM Calculée \n(à valider pour ce mois)" else HEADER_FONT
                )
                cell.number_format = "#,##0"

            elif col == "COMMENTAIRE":
                cell = ws_annexe_1.cell(row=start, column=element)
                cell.border = THIN_BORDER

            else:
                cell = ws_annexe_1.cell(row=start, column=element)
                apply_format_on_dmm_cell(cell)
                # Il faut également faire des mise à jour de la cellule droite
                cell = ws_annexe_1.cell(row=start, column=element + 1)
                apply_format_on_dmm_cell(cell, True)

    return ws_annexe_1, df_produit


def update_cmm_informations_on_sheet(
    wb_temp: Workbook,
    df_dmm_global: pd.DataFrame,
    df_dmm_histo: pd.DataFrame,
    df_cmm_global: pd.DataFrame,
    df_cmm_histo: pd.DataFrame,
    df_produit: pd.DataFrame,
    date_report: str,
):
    """
    Met à jour les informations de CMM (Consommation Moyenne Mensuelle) sur la feuille annexe 1 d'un classeur Excel.
    Args:
        wb_temp (Workbook): Le classeur Excel temporaire à mettre à jour.
        df_dmm_global (pd.DataFrame): DataFrame contenant les données globales de DMM (Demande Moyenne Mensuelle).
        df_dmm_histo (pd.DataFrame): DataFrame contenant les données historiques de DMM.
        df_cmm_global (pd.DataFrame): DataFrame contenant les données globales de CMM.
        df_cmm_histo (pd.DataFrame): DataFrame contenant les données historiques de CMM.
        df_produit (pd.DataFrame): DataFrame contenant les informations sur les produits.
        date_report (str): La date du rapport sous forme de chaîne de caractères.
    Returns:
        Workbook: Le classeur Excel mis à jour avec les informations de CMM.
    """
    ws_annexe_1, df_produit = update_dmm_informations_on_sheet(
        wb_temp,
        df_dmm_global,
        df_dmm_histo,
        df_produit,
        date_report,
    )

    ws_annexe_1["CW2"].value = "Détermination de la CMM au mois de " + month_year_str

    df_cmm_to_sheet = pd.pivot_table(
        df_cmm_global,
        values="cmm",
        index=["id_dim_produit_stock_track_pk"],
        columns="date_report",
    ).reset_index()

    df_cmm_to_sheet = df_cmm_to_sheet.rename(
        columns=lambda x: str(x).replace(" 00:00:00", "").lstrip() if " 00:00:00" in str(x) else x
    )
    df_cmm_to_sheet.columns.name = ""

    # Cette jointure pour garder l'ordre des données
    df_cmm_to_sheet = df_produit.merge(
        df_cmm_to_sheet, on="id_dim_produit_stock_track_pk", how="left"
    ).sort_values("CODE")[df_cmm_to_sheet.columns.tolist()]

    cols_df_dmm = df_cmm_to_sheet.columns.tolist()

    header_row = list(ws_annexe_1.iter_rows(min_row=4, max_row=4, min_col=64, max_col=99))[0]
    dico_cols = {}

    for cell in header_row:
        if cell.value is None or not hasattr(cell, "col_idx"):
            continue
        elif cell.value.strftime("%Y-%m-%d") in cols_df_dmm:
            dico_cols[cell.value.strftime("%Y-%m-%d")] = [
                cell.column_letter,
                cell.col_idx,
                cols_df_dmm.index(cell.value.strftime("%Y-%m-%d")),
                max(
                    [
                        len(val)
                        for val in df_cmm_to_sheet[cell.value.strftime("%Y-%m-%d")].astype(str)
                    ]
                ),
            ]
        else:
            dico_cols[cell.value.strftime("%Y-%m-%d")] = cell.col_idx

    # Gestion des autres colonnes renseignant sur la DMM du mois courant
    header_row = list(ws_annexe_1.iter_rows(min_row=3, max_row=3, min_col=101, max_col=104))[0]

    dico_cols.update({cell.value: cell.col_idx for cell in header_row})

    dico_formules = {
        "Nbre de mois de considérés": '=COUNTIFS(BL{0}:CU{0},"X")',
        "Consommations enregistrées sur les mois de considérés": '=SUM(IF(BM{0}="X",BL{0},0),IF(BO{0}="X",BN{0},0),IF(BQ{0}="X",BP{0},0),IF(BS{0}="X",BR{0},0),IF(BU{0}="X",BT{0},0),IF(BW{0}="X",BV{0},0),IF(BY{0}="X",BX{0},0),IF(CA{0}="X",BZ{0},0),IF(CC{0}="X",CB{0},0),IF(CE{0}="X",CD{0},0),IF(CG{0}="X",CF{0},0),IF(CI{0}="X",CH{0},0),IF(CK{0}="X",CJ{0},0),IF(CM{0}="X",CL{0},0),IF(CO{0}="X",CN{0},0),IF(CQ{0}="X",CP{0},0),IF(CS{0}="X",CR{0},0), IF(CU{0}="X",CT{0},0))',
        "CMM Calculée en fin du mois": "=IFERROR(CX{0}/CW{0},0)",
    }

    formula_cell_cv = "=ROUNDUP(IFERROR(VLOOKUP(A{0},StockParRegion!A:F,6,FALSE),0)/H{0},0)"  # "=ROUNDUP(IFERROR(VLOOKUP(A{0},StockParRegion!A:F,6,FALSE),0)/F{0},0)"
    # print(dico_cols)
    for start, row in enumerate(
        dataframe_to_rows(df_cmm_to_sheet, index=False, header=False), start=5
    ):
        # formule dans la barre de CS pour retrouver la valeur de CMM courante pour le mois
        cell = ws_annexe_1.cell(row=start, column=100, value=formula_cell_cv.format(start))
        cell.font = CS_FONT
        for col, element in dico_cols.items():
            if col == date_format.strftime("%Y-%m-%d"):
                column_element = element if not isinstance(element, list) else element[1]
                cell = ws_annexe_1.cell(row=start, column=column_element, value=f"=$CV{start}")
                apply_format_on_dmm_cell(cell)
                # Il faut également faire des mise à jour de la cellule droite
                cell = ws_annexe_1.cell(row=start, column=column_element + 1, value="X")
                apply_format_on_dmm_cell(cell, True)
            elif not isinstance(element, int):
                cell = ws_annexe_1.cell(row=start, column=element[1], value=row[element[2]])
                apply_format_on_dmm_cell(cell)
                # Gestion de la Céllule de droite !!!!
                # Ici il faut avoir le nombre de mois considérés sur les données courant
                # Recherche de la valeur pour le coching de la cellule suivante
                mois_considere = df_cmm_histo.loc[
                    (df_cmm_histo.id_dim_produit_stock_track_pk == row[0]), "date_report_prev"
                ].values
                mois_considere = [date.strftime("%Y-%m-%d") for date in mois_considere]

                val = (
                    "X"
                    if col
                    in mois_considere  # pd.date_range(end=date_format, periods=int(mois_considere), freq="MS")
                    else None
                )
                cell = ws_annexe_1.cell(row=start, column=element[1] + 1, value=val)
                apply_format_on_dmm_cell(cell, True)

            elif col in dico_formules:
                cell = ws_annexe_1.cell(
                    row=start, column=element, value=dico_formules.get(col).format(start)
                )
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGNMENT
                cell.font = BODY_FONT if col != "CMM Calculée en fin du mois" else HEADER_FONT
                cell.number_format = "#,##0"

            elif col == "COMMENTAIRE":
                cell = ws_annexe_1.cell(row=start, column=element)
                cell.border = THIN_BORDER

            else:
                cell = ws_annexe_1.cell(row=start, column=element)
                apply_format_on_dmm_cell(cell)
                # Il faut également faire des mise à jour de la cellule droite
                cell = ws_annexe_1.cell(row=start, column=element + 1)
                apply_format_on_dmm_cell(cell, True)

    return wb_temp


def update_sheet_annexe_1(
    wb_temp: Workbook, programme: str, schema_name: str, engine: Engine, date_report: str
):
    """
    Met à jour la feuille annexe 1 avec les informations CMM et DMM.
    Args:
        wb_temp (Workbook): Le classeur Excel temporaire à mettre à jour.
        programme (str): Le nom du programme pour lequel les données doivent être récupérées.
        schema_name (str): Le nom du schéma de la base de données.
        engine (Engine): L'objet moteur de la base de données pour exécuter les requêtes SQL.
        date_report (str): La date du rapport sous forme de chaîne de caractères.
    Returns:
        Workbook: Le classeur Excel mis à jour avec les informations CMM et DMM.
    """
    f_date_report = pd.to_datetime(date_report, format="%d/%m/%Y", errors="coerce").strftime(
        "%Y-%m-%d"
    )

    df_produit = pd.read_sql(
        QUERY_ANNEXE_1_PRODUIT.format(
            schema_name=schema_name,
            programme=programme,
            date_report=f_date_report,
        ),
        engine,
    )
    df_dmm_global = pd.read_sql(
        QUERY_ANNEXE_1_DMM_GLOBAL.format(schema_name=schema_name, programme=programme), engine
    )
    df_dmm_histo = pd.read_sql(
        QUERY_ANNEXE_1_DMM_HISTO.format(
            schema_name=schema_name,
            programme=programme,
            date_report=f_date_report,
        ),
        engine,
    )
    df_cmm_global = pd.read_sql(
        QUERY_ANNEXE_1_CMM_GLOBAL.format(schema_name=schema_name, programme=programme), engine
    )
    df_cmm_histo = pd.read_sql(
        QUERY_ANNEXE_1_CMM_HISTO.format(
            schema_name=schema_name,
            programme=programme,
            date_report=f_date_report,
        ),
        engine,
    )

    for data in (df_dmm_global, df_dmm_histo, df_cmm_global, df_cmm_histo):
        data["date_report"] = pd.to_datetime(data["date_report"])

    return update_cmm_informations_on_sheet(
        wb_temp,
        df_dmm_global,
        df_dmm_histo,
        df_cmm_global,
        df_cmm_histo,
        df_produit,
        date_report,
    )

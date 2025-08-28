from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle

# Constants and shared configurations
DICO_MOIS_FR = {
    "January": "Janvier",
    "February": "Février",
    "March": "Mars",
    "April": "Avril",
    "May": "Mai",
    "June": "Juin",
    "July": "Juillet",
    "August": "Août",
    "September": "Septembre",
    "October": "Octobre",
    "November": "Novembre",
    "December": "Décembre",
}

# Styles definitions
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

HEADER_FONT = Font(name="Arial Narrow", size=12, bold=True)
BODY_FONT = Font(name="Arial Narrow", size=12)
ALERT_FONT = Font(name="Arial Narrow", size=12, bold=True, color="FFC00000")
CS_FONT = Font(name="Arial Narrow", size=12, color="FFFFFF")

ALIGNMENT = Alignment(horizontal=None, vertical="center")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")

LIGHT_BLUE_FILL = PatternFill(start_color="FFDDEBF7", fill_type="solid")

DATE_STYLE = NamedStyle(
    name="date_style",
    number_format="DD/MM/YYYY",
    border=THIN_BORDER,
    font=BODY_FONT,
    alignment=CENTER_ALIGNMENT,
)

DICO_RULES_ANNEXE_1 = {
    "rule_type_produit": CellIsRule(
        operator="equal",
        formula=['"Traceur"'],
        fill=PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"),
        font=Font(name="Arial Narrow", size=12),
    ),
    "rule_less_than_zero": Rule(
        type="cellIs",
        operator="lessThan",
        formula=["0"],
        dxf=DifferentialStyle(font=Font(name="Arial Narrow", size=11, bold=True, color="FFFF0000")),
        priority=2,
    ),
    "rule_greater_than_zero": Rule(
        type="cellIs",
        operator="greaterThan",
        formula=["0"],
        dxf=DifferentialStyle(font=Font(name="Arial Narrow", size=11, bold=True, color="FF0033CC")),
        priority=3,
    ),
    "rule_equal_zero": Rule(
        type="cellIs",
        operator="equal",
        formula=["0"],
        dxf=DifferentialStyle(
            font=Font(name="Arial Narrow", size=11, bold=False, color="FFA6A6A6")
        ),
        priority=1,
    ),
}


DICO_RULES_ANNEXE_2 = {
    "Stock Dormant": DifferentialStyle(
        fill=PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid"),
        font=Font(name="Arial Narrow", size=11, bold=True),
    ),
    "SurStock": DifferentialStyle(
        fill=PatternFill(start_color="FF5B9BD5", end_color="FF5B9BD5", fill_type="solid"),
        font=Font(name="Arial Narrow", size=10, bold=True),
    ),
    "Rupture": DifferentialStyle(
        fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
        font=Font(name="Arial Narrow", size=10, bold=True),
    ),
    "Bien Stocké": DifferentialStyle(
        fill=PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid"),
        font=Font(name="Arial Narrow", size=11, bold=True),
    ),
    "Sous-Stock": DifferentialStyle(
        fill=PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),
        font=Font(name="Arial Narrow", size=10, bold=True),
    ),
}

QUERY_ANNEXE_1_PRODUIT = """
SELECT prod.*, st.stock_theorique_mois_precedent
FROM {schema_name}.stock_track st
INNER JOIN {schema_name}.dim_produit_stock_track prod ON st.id_dim_produit_stock_track_fk = prod.id_dim_produit_stock_track_pk
WHERE prod.programme='{programme}' AND st.date_report='{date_report}'
"""

QUERY_ANNEXE_1_DMM_GLOBAL = """
SELECT prod.*, st_dmm.* 
FROM {schema_name}.stock_track_dmm st_dmm
INNER JOIN {schema_name}.dim_produit_stock_track prod ON st_dmm.id_dim_produit_stock_track_fk = prod.id_dim_produit_stock_track_pk
WHERE prod.programme='{programme}'
"""

QUERY_ANNEXE_1_DMM_HISTO = """
SELECT prod.*, st_dmm_histo.date_report, st_dmm_histo.date_report_prev, st_dmm_histo.dmm
FROM {schema_name}.stock_track_dmm_histo st_dmm_histo
INNER JOIN {schema_name}.dim_produit_stock_track prod ON st_dmm_histo.id_dim_produit_stock_track_fk = prod.id_dim_produit_stock_track_pk
WHERE prod.programme='{programme}' AND st_dmm_histo.date_report='{date_report}'
"""
QUERY_ANNEXE_1_CMM_GLOBAL = """
SELECT prod.*, st_cmm.* 
FROM {schema_name}.stock_track_cmm st_cmm
INNER JOIN {schema_name}.dim_produit_stock_track prod ON st_cmm.id_dim_produit_stock_track_fk = prod.id_dim_produit_stock_track_pk
WHERE prod.programme='{programme}'
"""
QUERY_ANNEXE_1_CMM_HISTO = """
SELECT prod.*, st_cmm_histo.date_report, st_cmm_histo.date_report_prev, st_cmm_histo.cmm
FROM {schema_name}.stock_track_cmm_histo st_cmm_histo
INNER JOIN {schema_name}.dim_produit_stock_track prod ON st_cmm_histo.id_dim_produit_stock_track_fk = prod.id_dim_produit_stock_track_pk
WHERE prod.programme='{programme}' AND st_cmm_histo.date_report='{date_report}'
"""

DICO_FORMULES_ANNEXE_2 = {
    1: "='Annexe 1 - Consolidation'!A{0}",  # CODE
    2: "='Annexe 1 - Consolidation'!B{0}",  # Ancien code
    3: "='Annexe 1 - Consolidation'!E{0}",  # CATEGORIE
    4: "='Annexe 1 - Consolidation'!D{0}",  # DESIGNATION DU PRODUIT
    5: "='Annexe 1 - Consolidation'!C{0}",  # Type
    6: "='Annexe 1 - Consolidation'!H{0}",  # Facteur de conversion \n(De la centrale à la périphérie)
    7: "='Annexe 1 - Consolidation'!F{0}",  # Unité niv Central
    8: "='Annexe 1 - Consolidation'!G{0}",  # Unité Niveau Périphérique
    9: "='Annexe 1 - Consolidation'!O{0}",  # SDU (niveau central)
    10: "='Annexe 1 - Consolidation'!BI{0}",  # DMM (niveau central)
    11: '=IF(I{0}=0,0,IFERROR(I{0}/J{0},"ND"))',  # MSD (niveau central)
    12: '=IF(I{0}=0,"Rupture",IF(J{0}=0,"Stock dormant",IF(K{0}<3,"Sous-Stock",IF(K{0}>8,"SurStock","Bien Stocké"))))',  # STATUT (niveau central)
    13: "=IFERROR(ROUNDUP(VLOOKUP($A{0}, StockParRegion!$A:$J,6,FALSE)/F{0}, 0), 0)",  # CONSO (niveau décentralisé)
    14: "=IFERROR(ROUNDUP(VLOOKUP($A{0}, StockParRegion!$A:$J,7,FALSE)/F{0}, 0), 0)",  # SDU (niveau décentralisé)
    15: "=IFERROR(ROUNDUP(VLOOKUP($A{0},StockParRegion!$A:$J,8,FALSE)/F{0}, 0), 0)",  # CMM (niveau décentralisé)
    16: "=IFERROR(VLOOKUP($A{0},StockParRegion!$A:$J,9,FALSE),0)",  # MSD (niveau décentralisé)
    # STATUT (niveau décentralisé)
    17: '=IF(N{0}=0,"Rupture",IF(O{0}=0,"Stock dormant",IF(P{0}<2,"Sous-Stock",IF(P{0}>4,"SurStock","Bien Stocké"))))',
    # Nombre de Sites en Rupture (niveau décentralisé)
    18: "=COUNTIFS('Etat de stock Periph'!A:A,A{0},'Etat de stock Periph'!Y:Y,\"RUPTURE\")",
    19: "=I{0}+N{0}",  # SDU (niveau National)
    20: "=O{0}",  # CMM (niveau National)
    21: '=IF(S{0}=0,0,IFERROR(S{0}/T{0},"ND"))',  # MSD (niveau National)
    22: '=IF(S{0}=0,"Rupture",IF(T{0}=0,"Stock dormant",IF(U{0}<5, "Sous-Stock", IF(U{0}>12,"SurStock","Bien Stocké"))))',  # STATUT (niveau National)
    # ---> Date de Péremption la plus proche (BRUTE)!!! MINIFS # MIN.SI.ENS
    23: "=MINIFS('Stock detaille'!D:D,'Stock detaille'!G:G,\">0\",'Stock detaille'!A:A, A{0})",  # Date de Péremption la plus proche (BRUTE)
    24: '=IF(W{0}>0,W{0},"")',  # Date de Péremption la plus proche
    25: "=IF(X{0}=\"\",\"\",SUMIFS('Stock detaille'!G:G,'Stock detaille'!A:A,A{0},'Stock detaille'!D:D, X{0}))",  # Quantité correspondante
    26: '=IF(Y{0}="","",IFERROR(Y{0}/J{0},"NA"))',  # MSD correspondant
    27: '=IFERROR(IF(Z{0}="","",(X{0}-TODAY()-90)/30.438),"")',  # Durée d'utilisation à la NPSP (mois)
    28: '=IFERROR(X{0}-TODAY(),"")',  # Days until expiration
    29: "=IFERROR(SUMIFS('Stock detaille'!G:G,'Stock detaille'!A:A,A{0},'Stock detaille'!K:K,\"RED\")/J{0},\"\")",  # Moins de 6 mois (RED)
    30: "=IFERROR(SUMIFS('Stock detaille'!G:G,'Stock detaille'!A:A,A{0},'Stock detaille'!K:K,\"ORANGE\")/J{0},\"\")",  # Entre 6 et 12 mois (ORANGE)
    31: "=IFERROR(SUMIFS('Stock detaille'!G:G,'Stock detaille'!A:A,A{0},'Stock detaille'!K:K,\"GREEN\")/J{0},\"\")",  # Plus de 12 mois (GREEN)
    32: "=IF(AK{0}=0,\"\",SUMIFS('Plan d''appro'!J:J,'Plan d''appro'!A:A,A{0},'Plan d''appro'!K:K, AK{0}))",  # Qtité attendue
    33: "=IFERROR(AF{0}/J{0},0)",  # MSD attendu
    34: '=SUMIFS(Receptions!H:H,Receptions!C:C,A{0},Receptions!J:J,"<>ok")',  # Qtité réceptionnés non en Stock
    35: "=IFERROR(AH{0}/J{0},0)",  # MSD reçu
    # 36: None,
    36: "=IFERROR(INDEX('Plan d''appro'!F:F,MATCH(A{0}&\"_\"&AK{0},'Plan d''appro'!T:T,0),1),\"\")",  # Financement
    # ---> Date Probable de Livraison !!! MINIFS
    37: "=IFERROR(MINIFS('Plan d''appro'!K:K,'Plan d''appro'!A:A,A{0},'Plan d''appro'!K:K,\">\"&$AK$1),\"\")",
    # 39: None,
    # ---> Date effective de livraison !!! MAXIFS # MAX.SI.ENS
    38: '=MAXIFS(Receptions!F:F,Receptions!C:C,A{0}, Receptions!J:J, "ok")',
    # ---> Delivery status
    39: "=IFERROR(INDEX('Plan d''appro'!G:G,MATCH(A{0}&\"_\"&AK{0},'Plan d''appro'!T:T,0),1),\"\")",
}

DICO_FORMULES_PREVISION = {
    "h": '=IF($E{0}="", "", VLOOKUP($E{0},\'Annexe 2 - Suivi des Stocks\'!$A:$V, 9, FALSE))',  # STOCK CENTRAL
    "i": '=IF($E{0}="", "", VLOOKUP($E{0},\'Annexe 2 - Suivi des Stocks\'!$A:$V, 10,FALSE))',  # DMM
    "j": '=IF(H{0}=0, 0, IFERROR(H{0}/I{0},"ND"))',  # mois_courant
    "k": "=IFERROR(MAX(0,J{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,K$6)/$I{0},\"ND\")",
    "l": "=IFERROR(MAX(0,K{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,L$6)/$I{0},\"ND\")",
    "m": "=IFERROR(MAX(0,L{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,M$6)/$I{0},\"ND\")",
    "n": "=IFERROR(MAX(0,M{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,N$6)/$I{0},\"ND\")",
    "o": "=IFERROR(MAX(0,N{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,O$6)/$I{0},\"ND\")",
    "p": "=IFERROR(MAX(0,O{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,P$6)/$I{0},\"ND\")",
    "q": "=IFERROR(MAX(0,P{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,Q$6)/$I{0},\"ND\")",
    "r": "=IFERROR(MAX(0,Q{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,R$6)/$I{0},\"ND\")",
    "s": "=IFERROR(MAX(0,R{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,S$6)/$I{0},\"ND\")",
    "t": "=IFERROR(MAX(0,S{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,T$6)/$I{0},\"ND\")",
    "u": "=IFERROR(MAX(0,T{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,U$6)/$I{0},\"ND\")",
    "v": "=IFERROR(MAX(0,U{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,V$6)/$I{0},\"ND\")",
    "x": "=E{0}",
    "y": "=F{0}",
    "z": "=G{0}",
    "aa": '=IF($E{0}="", "", VLOOKUP($E{0},\'Annexe 2 - Suivi des Stocks\'!A:V,19,FALSE))',  # Stock National
    "ab": '=IF($E{0}="", "", VLOOKUP($E{0},\'Annexe 2 - Suivi des Stocks\'!A:V,20,FALSE))',  # CMM
    "ac": '=IF(AA{0}=0, 0, IFERROR(AA{0}/AB{0},"ND"))',  # mois_courant
    "ad": "=IFERROR(MAX(0,AC{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AD$6)/$AB{0},\"ND\")",
    "ae": "=IFERROR(MAX(0,AD{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AE$6)/$AB{0},\"ND\")",
    "af": "=IFERROR(MAX(0,AE{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AF$6)/$AB{0},\"ND\")",
    "ag": "=IFERROR(MAX(0,AF{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AG$6)/$AB{0},\"ND\")",
    "ah": "=IFERROR(MAX(0,AG{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AH$6)/$AB{0},\"ND\")",
    "ai": "=IFERROR(MAX(0,AH{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AI$6)/$AB{0},\"ND\")",
    "aj": "=IFERROR(MAX(0,AI{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AJ$6)/$AB{0},\"ND\")",
    "ak": "=IFERROR(MAX(0,AJ{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AK$6)/$AB{0},\"ND\")",
    "al": "=IFERROR(MAX(0,AK{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AL$6)/$AB{0},\"ND\")",
    "am": "=IFERROR(MAX(0,AL{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AM$6)/$AB{0},\"ND\")",
    "an": "=IFERROR(MAX(0,AM{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AN$6)/$AB{0},\"ND\")",
    "ao": "=IFERROR(MAX(0,AN{0}-1)+SUMIFS('Plan d''appro'!$J:$J,'Plan d''appro'!$A:$A,$E{0},'Plan d''appro'!$P:$P,AO$6)/$AB{0},\"ND\")",
    "aq": "=E{0}",
    "ar": "=G{0}",
    "as": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&AS$6, 'Stock detaille'!$D:$D, \"<\"&AT$6)",
    "at": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&AT$6, 'Stock detaille'!$D:$D, \"<\"&AU$6)",
    "au": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&AU$6, 'Stock detaille'!$D:$D, \"<\"&AV$6)",
    "av": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&AV$6, 'Stock detaille'!$D:$D, \"<\"&AW$6)",
    "aw": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&AW$6, 'Stock detaille'!$D:$D, \"<\"&AX$6)",
    "ax": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&AX$6, 'Stock detaille'!$D:$D, \"<\"&AY$6)",
    "ay": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&AY$6, 'Stock detaille'!$D:$D, \"<\"&AZ$6)",
    "az": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&AZ$6, 'Stock detaille'!$D:$D, \"<\"&BA$6)",
    "ba": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&BA$6, 'Stock detaille'!$D:$D, \"<\"&BB$6)",
    "bb": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&BB$6, 'Stock detaille'!$D:$D, \"<\"&BC$6)",
    "bc": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&BC$6, 'Stock detaille'!$D:$D, \"<\"&BD$6)",
    "bd": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&BD$6, 'Stock detaille'!$D:$D, \"<\"&BE$6)",
    "be": "=SUMIFS('Stock detaille'!$G:$G, 'Stock detaille'!$A:$A, $AQ{0}, 'Stock detaille'!$D:$D, \">=\"&BE$6, 'Stock detaille'!$D:$D, \"<\"&BF$6)",
    "bg": "=E{0}",
    "bh": "=G{0}",
    "bi": '=IF(AS{0}=0, 0, IFERROR(AS{0}/I{0},"ND"))',
    "bj": '=IF(AT{0}=0, 0, IFERROR(AT{0}/I{0},"ND"))',
    "bk": '=IF(AU{0}=0, 0, IFERROR(AU{0}/I{0},"ND"))',
    "bl": '=IF(AV{0}=0, 0, IFERROR(AV{0}/I{0},"ND"))',
    "bm": '=IF(AW{0}=0, 0, IFERROR(AW{0}/I{0},"ND"))',
    "bn": '=IF(AX{0}=0, 0, IFERROR(AX{0}/I{0},"ND"))',
    "bo": '=IF(AY{0}=0, 0, IFERROR(AY{0}/I{0},"ND"))',
    "bp": '=IF(AZ{0}=0, 0, IFERROR(AZ{0}/I{0},"ND"))',
    "bq": '=IF(BA{0}=0, 0, IFERROR(BA{0}/I{0},"ND"))',
    "br": '=IF(BB{0}=0, 0, IFERROR(BB{0}/I{0},"ND"))',
    "bs": '=IF(BC{0}=0, 0, IFERROR(BC{0}/I{0},"ND"))',
    "bt": '=IF(BD{0}=0, 0, IFERROR(BD{0}/I{0},"ND"))',
    "bu": '=IF(BE{0}=0, 0, IFERROR(BE{0}/I{0},"ND"))',
    "bw": "=E{0}",
    "bx": "=G{0}",
    "by": "=IF(AS{0}=0, 0, IFERROR(AS{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "bz": "=IF(AT{0}=0, 0, IFERROR(AT{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "ca": "=IF(AU{0}=0, 0, IFERROR(AU{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "cb": "=IF(AV{0}=0, 0, IFERROR(AV{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "cc": "=IF(AW{0}=0, 0, IFERROR(AW{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "cd": "=IF(AX{0}=0, 0, IFERROR(AX{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "ce": "=IF(AY{0}=0, 0, IFERROR(AY{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "cf": "=IF(AZ{0}=0, 0, IFERROR(AZ{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "cg": "=IF(BA{0}=0, 0, IFERROR(BA{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "ch": "=IF(BB{0}=0, 0, IFERROR(BB{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "ci": "=IF(BC{0}=0, 0, IFERROR(BC{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "cj": "=IF(BD{0}=0, 0, IFERROR(BD{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
    "ck": "=IF(BE{0}=0, 0, IFERROR(BE{0}*VLOOKUP($BW{0}, 'Plan d''appro'!$A:$S, 19, FALSE),\"ND\"))",
}


DICO_RULES_PREVISION = {
    # Gestion de mise en forme conditionnelle
    "rule_equal_zero": Rule(
        type="cellIs",
        operator="equal",
        formula=["0"],
        dxf=DifferentialStyle(
            fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        ),
        # priority=1,
    ),
    "rule_less_than_third": Rule(
        type="cellIs",
        operator="lessThan",
        formula=["3"],
        dxf=DifferentialStyle(
            fill=PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
        ),
        # priority=2,
    ),
    "rule_between_third_and_eight": Rule(
        type="cellIs",
        operator="lessThanOrEqual",
        formula=["8"],
        dxf=DifferentialStyle(
            fill=PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
        ),
        # priority=3,
    ),
    "rule_greater_than_eight": Rule(
        type="cellIs",
        operator="greaterThan",
        formula=["8"],
        dxf=DifferentialStyle(
            fill=PatternFill(start_color="FF8EAADB", end_color="FF8EAADB", fill_type="solid")
        ),
        # priority=4,
    ),
    "rule_equal_nd": Rule(
        type="cellIs",
        operator="equal",
        formula=['"ND"'],
        dxf=DifferentialStyle(
            fill=PatternFill(start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")
        ),
        # priority=9,
    ),
    "rule_less_than_five": Rule(
        type="cellIs",
        operator="lessThan",
        formula=["5"],
        dxf=DifferentialStyle(
            fill=PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
        ),
        # priority=6,
    ),
    "rule_between_five_and_twelve": Rule(
        type="cellIs",
        operator="lessThanOrEqual",
        formula=["12"],
        dxf=DifferentialStyle(
            fill=PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
        ),
        # priority=7,
    ),
    "rule_greater_than_twelve": Rule(
        type="cellIs",
        operator="greaterThan",
        formula=["12"],
        dxf=DifferentialStyle(
            fill=PatternFill(start_color="FF8EAADB", end_color="FF8EAADB", fill_type="solid")
        ),
        # priority=8,
    ),
    "rule_greater_than_zero": Rule(
        type="cellIs",
        operator="greaterThan",
        formula=["0"],
        dxf=DifferentialStyle(
            fill=PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
        ),
        # priority=5,
    ),
    "rule_equal_empty": Rule(
        type="cellIs",
        operator="equal",
        formula=[""],
        dxf=DifferentialStyle(
            fill=PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
        ),
        # priority=1,
    ),
}

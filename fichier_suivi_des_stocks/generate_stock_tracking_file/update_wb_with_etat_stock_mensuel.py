from compute_indicators.utils import check_if_sheet_name_in_file
from efc.interfaces.iopenpyxl import OpenpyxlInterface
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle

from .utils import find_best_match, get_current_variable, has_formula


def update_data_on_sheet(wb_base, ws_base, ws_temp, sheet_name, programme, date_report, max_row):
    """
    Sert à actualiser les donnée des feuilles en utilisant la source de d'origine

    Args:
        wb_base (workbook): workbook etat stock mensuel
        ws_temp (workbook): worksheet du template
        sheet_name (str): le nom de la feuille
        programme (str): le nom du programme
        date_report (str): date de conception du rapport
        max_row (int): à partir de quelle ligne dois-t-on commencer les itérations
    """

    dico_cols = {}

    if sheet_name == "Etat de stock":
        ws_temp["B1"].value = ws_temp["B1"].value.replace("PNLP", programme.upper())
        ws_temp["B4"].value = ws_temp["B4"].value.replace("JUILLET 2024", month_year_str)
        if programme.upper() == "PNLT":
            ws_temp["B4"].value = ws_temp["B4"].value.replace(
                "3 mois / Max: 8 mois", "8 mois / Max: 12 mois"
            )

        ws_temp["L5"].value = "Stock théorique fin " + month_year_str.capitalize()
        # ws_temp['K5'].value = 'Nombre de jour de rupture en '+ prev_month_year_str
        ws_temp.title = "Etat de stock " + programme.upper()

    for row_base, row_temp in zip(
        ws_base.iter_rows(min_row=max_row, max_row=max_row),
        ws_temp.iter_rows(min_row=max_row, max_row=max_row),
    ):
        values = [cell.value for cell in row_base if cell.value is not None]
        for cell in row_temp:
            match_index = (
                find_best_match(cell.value, values)
                if cell.value.strip() != "Quantité livrée"
                else find_best_match("Quantité livrée", values)
                or find_best_match("Qté livrée", values)
            )
            if match_index is not None:
                dico_cols[match_index] = cell.col_idx

    interface = OpenpyxlInterface(wb=wb_base, use_cache=True)
    interface.clear_cache()

    font = Font(name="Calibri", size=11)
    alignment_one = Alignment(horizontal="center", vertical="center")
    alignment_two = Alignment(horizontal=None, vertical="center")
    fill = PatternFill(start_color="FFA8D08D", fill_type="solid")

    max_column = 9 if sheet_name == "Stock detaille" else None

    # start = max_row+1
    # print(sheet_name)
    for start, row in enumerate(
        ws_base.iter_rows(min_row=max_row + 1, max_col=max_column), start=max_row + 1
    ):
        for cell in row:
            if hasattr(cell, "col_idx") and dico_cols.get(cell.col_idx) is None:
                pass
            elif hasattr(cell, "col_idx") and has_formula(cell):
                result = interface.calc_cell(cell.coordinate, ws_base.title)
                new_cell = ws_temp.cell(row=cell.row, column=dico_cols[cell.col_idx], value=result)
            elif hasattr(cell, "col_idx"):
                new_cell = ws_temp.cell(
                    row=cell.row, column=dico_cols[cell.col_idx], value=cell.value
                )
            else:
                pass

            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = (
                    "DD/MM/YYYY"
                    if (sheet_name == "Distribution" and new_cell.col_idx in (4, 2))
                    or (sheet_name == "Receptions" and new_cell.col_idx == 9)
                    else cell.number_format
                )
                new_cell.alignment = cell.alignment.copy()

            # if (sheet_name == "Distribution" and new_cell.col_idx in (4, 2)) or (
            #     sheet_name == "Receptions" and new_cell.col_idx == 9
            # ):
            #     new_cell.number_format = "DD/MM/YYYY"

        # start+=1
        if sheet_name == "Stock detaille":
            ws_temp[f"J{start}"] = '=IFERROR(D{0}-TODAY(),"")'.format(start)
            ws_temp[f"K{start}"] = (
                '=IF(J{0}<180,"RED", IF(AND(J{0}>=180,J{0}<=365),"ORANGE","GREEN"))'.format(start)
            )
            ws_temp[f"J{start}"].font = ws_temp[f"K{start}"].font = font
            ws_temp[f"J{start}"].fill = ws_temp[f"K{start}"].fill = fill
            ws_temp[f"J{start}"].alignment = alignment_one
            ws_temp[f"K{start}"].alignment = alignment_two

        if sheet_name == "Receptions":
            ws_temp[f"J{start}"] = (
                '=IFERROR(IF(AND(YEAR(I{index})={year}, MONTH(I{index})={month}), "ok", "skip"), "skip")'.format(
                    index=start, year=date_report.year, month=date_report.month
                )
            )
            ws_temp[f"J{start}"].font = font

    # Gestion des céllules fusionnées
    for merged_cell in ws_base.merged_cells.ranges:
        if merged_cell.max_row >= max_row + 1:
            min_col_, min_row_, max_col_, max_row_ = merged_cell.bounds

            for row in range(min_row_, max_row_ + 1):
                for col in range(min_col_, max_col_ + 1):
                    cell = ws_base.cell(row=row, column=col)
                    new_cell = ws_temp.cell(row=row, column=col)

                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = cell.alignment.copy()

            ws_temp.merge_cells(str(merged_cell))

    if sheet_name == "Etat de stock":
        dico_dxf = {
            "Disponible à l'Agence": DifferentialStyle(
                fill=PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid"),
                font=Font(name="Tahoma", size=10, bold=True, color="FF000000", italic=True),
            ),
            "Rupture à l'Agence": DifferentialStyle(
                fill=PatternFill(start_color="FFFF6600", end_color="FFFF6600", fill_type="solid"),
                font=Font(name="Tahoma", size=10, bold=True, color="FF000000", italic=True),
            ),
            "Produit disponible": DifferentialStyle(
                fill=PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid"),
                font=Font(name="Tahoma", size=10, bold=True, color="FF000000", italic=True),
            ),
            "Rupture": DifferentialStyle(
                fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
                font=Font(name="Tahoma", size=10, bold=True, color="FFFFFFFF", italic=True),
            ),
            "ND": DifferentialStyle(
                fill=PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"),
                font=Font(name="Tahoma", size=10, bold=True, color="FF000000", italic=True),
            ),
            "Bon": DifferentialStyle(
                fill=PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid"),
                font=Font(name="Tahoma", size=10, bold=True, color="FF000000", italic=True),
            ),
            "Sous-stock": DifferentialStyle(
                fill=PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),
                font=Font(name="Tahoma", size=10, bold=True, color="FF000000", italic=True),
            ),
            "Surstock": DifferentialStyle(
                fill=PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid"),
                font=Font(name="Tahoma", size=10, bold=True, color="FF000000", italic=True),
            ),
        }

        f_rule_etat_stock = (
            'NOT(ISERROR(SEARCH("{etat_stock}", {col_letter_etat_stock}{index_start})))'
        )
        for etat_stock in dico_dxf:
            rule = Rule(
                type="containsText",
                operator="containsText",
                dxf=dico_dxf[etat_stock],
                text=etat_stock,
                formula=[
                    f_rule_etat_stock.format(
                        etat_stock=etat_stock, col_letter_etat_stock="N", index_start=6
                    )
                ],
            )
            ws_temp.conditional_formatting.add(f"N6:N{start}", rule)


def update_sheets_etat_mensuel(wb_base, wb_temp, programme, date_report):
    """
    Fonction générale utilisée pour effectuer la mise à jour des données depuis le fichier etat de stock mensuel

    Args:
        wb_base (workbook): workbook base
        wb_temp (workbook): workbook template
        programme (str): le nom du programme
        date_report (str): date de conception du rapport
    """
    global date_format, month_year_str, prev_month_year_str

    date_format, month_year_str, prev_month_year_str = get_current_variable(date_report)

    dico_sheet_names = {
        "Etat de stock": 5,
        "Stock detaille": 1,
        "Receptions": 1,
        "Distribution": 1,
        "Produits en transfert": 3,
        "PPI": 3,
        "Prelèvement": 3,
    }
    sheet_names_ws_base = wb_base.sheetnames
    sheet_names_ws_temp = wb_temp.sheetnames

    for sheet_name, max_row in dico_sheet_names.items():
        _sheet_name = check_if_sheet_name_in_file(sheet_name, sheet_names_ws_base)
        ws_base = wb_base[_sheet_name]

        _sheet_name = check_if_sheet_name_in_file(sheet_name, sheet_names_ws_temp)
        ws_temp = wb_temp[_sheet_name]

        update_data_on_sheet(wb_base, ws_base, ws_temp, sheet_name, programme, date_format, max_row)

        ws_temp.sheet_state = "hidden"

    return wb_temp

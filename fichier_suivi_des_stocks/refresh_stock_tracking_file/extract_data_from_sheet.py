import math

import numpy as np
import pandas as pd
from compute_indicators.utils import check_if_sheet_name_in_file
from efc.interfaces.iopenpyxl import OpenpyxlInterface  # type: ignore
from generate_stock_tracking_file.utils import has_formula
from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.utils import column_index_from_string

from .constants import COLUMNS_NAME_ETAT_STOCK, DICO_COLUMNS


def get_dmm_dataframes(
    df_etat_stock: pd.DataFrame, src_wb: Workbook, sheetnames: list[str], date_report: str
) -> tuple[pd.DataFrame]:
    """Extract DMM dataframes from the annex sheet.

    Args:
        df_etat_stock (pd.DataFrame): DataFrame containing stock data.
        src_wb (Workbook): Source workbook containing the annex sheet.
        sheetnames (list[str]): List of sheet names in the workbook.
        date_report (str): Date of the report in 'YYYY-MM-DD' format.

    Returns:
        tuple[pd.DataFrame]: Tuple containing two DataFrames:
            - df_stock_track_dmm: DataFrame with DMM stock tracking data.
            - df_stock_track_dmm_histo: DataFrame with historical DMM data.
    """
    sheet_annexe_1 = check_if_sheet_name_in_file("Annexe 1 - Consolidation", sheetnames)
    interface = OpenpyxlInterface(wb=src_wb, use_cache=True)
    interface.clear_cache()

    data_list = []
    for start, row in enumerate(
        src_wb[sheet_annexe_1].iter_rows(
            min_row=4,
            min_col=column_index_from_string("V"),
            max_col=column_index_from_string("BE"),
        ),
        start=1,
    ):
        data = []
        for cell in row:
            if has_formula(cell):
                result = interface.calc_cell(cell.coordinate, sheet_annexe_1)
                data.append(result)
            else:
                data.append(cell.value)
        if start == 1:
            new_data = []
            unnamed_counter = 22  # Colonne de début des DMM
            for value in data:
                if value is None:
                    new_data.append(f"Unnamed: {unnamed_counter}")
                    unnamed_counter += 23
                else:
                    new_data.append(value)
            data = new_data
        data_list.append(data)

    df_dmm = pd.concat(
        [df_etat_stock[["code_produit"]], pd.DataFrame(data_list[1:], columns=data_list[0])],
        axis=1,
    )

    df_stock_track_dmm = df_dmm[[col for col in df_dmm.columns if "Unnamed" not in str(col)]]

    df_stock_track_dmm = pd.melt(
        df_stock_track_dmm, id_vars="code_produit", var_name="date_report", value_name="dmm"
    )

    df_stock_track_dmm["date_report"] = (
        df_stock_track_dmm["date_report"]
        .apply(lambda x: pd.to_datetime(str(x)[:10], format="%Y-%m-%d"))
        .astype("<M8[ns]")
    )

    interface.clear_cache()
    data_list = []
    for row in src_wb[sheet_annexe_1].iter_rows(
        min_row=3,
        min_col=column_index_from_string("BG"),
        max_col=column_index_from_string("BJ"),
    ):
        data = []
        for cell in row:
            if has_formula(cell):
                result = interface.calc_cell(cell.coordinate, sheet_annexe_1)
                data.append(result)
            else:
                data.append(cell.value)
        data_list.append(data)

    data_list = [row for row in data_list if any(row)]

    df_dmm_curent_month = pd.concat(
        [
            df_etat_stock[["code_produit"]],
            pd.DataFrame(data_list[1:], columns=data_list[0]),
        ],
        axis=1,
    ).dropna(how="all")

    df_dmm_curent_month["date_report"] = pd.to_datetime(date_report, format="%Y-%m-%d")
    df_dmm_curent_month.columns = df_dmm_curent_month.columns.str.replace("\n", " ")
    df_dmm_curent_month.rename(
        columns={
            "Nbre de mois de considérés": "nbre_mois_consideres",
            "Distributions enregistrées sur les mois de considérés": "distributions_mois_consideres",
            "DMM Calculée  (à valider pour ce mois)": "dmm_calculee",
            "COMMENTAIRE": "commentaire",
        },
        inplace=True,
    )

    assert (
        df_stock_track_dmm.merge(
            df_dmm_curent_month, how="left", on=["code_produit", "date_report"]
        ).shape[0]
        == df_stock_track_dmm.shape[0]
    )

    df_stock_track_dmm = df_stock_track_dmm.merge(
        df_dmm_curent_month, how="left", on=["code_produit", "date_report"]
    )

    df_stock_track_dmm = df_stock_track_dmm.loc[df_stock_track_dmm["date_report"] == date_report]

    cols_df_dmm = df_dmm.columns.to_list()

    mapping = {
        col: cols_df_dmm[i - 1]
        for i, col in enumerate(cols_df_dmm)
        if "Unnamed" in str(col) and i > 0
    }

    df = pd.melt(
        df_dmm,
        id_vars=["code_produit"],
        value_vars=[col for col in df_dmm.columns if not pd.isna(col)],
        var_name="date_report",
    )

    df["date_report"] = df["date_report"].replace(mapping)  # .map(mapping)

    df_stock_track_dmm_histo = (
        df.loc[df["value"] == "X"]
        .drop(columns="value")
        .merge(df.loc[df["value"] != "X"], on=["code_produit", "date_report"])
        .rename(columns={"date_report": "date_report_prev"})
        .sort_values(
            ["code_produit", "date_report_prev"], ascending=[True, True], ignore_index=True
        )
    )

    df_stock_track_dmm_histo = df_stock_track_dmm_histo.rename(columns={"value": "dmm"})

    df_stock_track_dmm_histo["date_report"] = pd.to_datetime(date_report, format="%Y-%m-%d")

    return df_stock_track_dmm.round(0), df_stock_track_dmm_histo.round(0)


def get_cmm_dataframes(
    df_etat_stock: pd.DataFrame,
    df_stock_prog_nat: pd.DataFrame,
    src_wb: Workbook,
    sheetnames: list[str],
    date_report: str,
) -> tuple[pd.DataFrame]:
    """Extract CMM dataframes from the annex sheet.
    Args:
        df_etat_stock (pd.DataFrame): DataFrame containing stock data.
        df_stock_prog_nat (pd.DataFrame): DataFrame containing stock program data.
        src_wb (Workbook): Source workbook containing the annex sheet.
        sheetnames (list[str]): List of sheet names in the workbook.
        date_report (str): Date of the report in 'YYYY-MM-DD' format.
    Returns:
        tuple[pd.DataFrame]: Tuple containing two DataFrames:
            - df_stock_track_cmm: DataFrame with CMM stock tracking data.
            - df_stock_track_cmm_histo: DataFrame with historical CMM data.
    """
    sheet_annexe_1 = check_if_sheet_name_in_file("Annexe 1 - Consolidation", sheetnames)
    interface = OpenpyxlInterface(wb=src_wb, use_cache=True)
    interface.clear_cache()

    header_row = list(
        src_wb[sheet_annexe_1].iter_rows(
            min_row=4,
            max_row=4,
            min_col=column_index_from_string("BL"),
            max_col=column_index_from_string("CU"),
        )
    )[0]

    dico_cols = {}
    for cell in header_row:
        if not isinstance(cell, MergedCell):
            dico_cols[str(cell.value)[:10]] = cell.column

    for row in src_wb[sheet_annexe_1].iter_rows(
        min_row=5, min_col=dico_cols[date_report], max_col=dico_cols[date_report]
    ):
        for cell in row:
            if has_formula(cell):
                code_produit, facteur_conversion = (
                    src_wb[sheet_annexe_1].cell(cell.row, 1).value,
                    src_wb[sheet_annexe_1].cell(cell.row, 8).value,
                )
                df = df_stock_prog_nat.loc[df_stock_prog_nat["Code_produit"] == int(code_produit)]
                if not df.empty:
                    value = (
                        math.ceil(df.CONSO.sum() / int(facteur_conversion))
                        if not pd.isna(facteur_conversion) and facteur_conversion != 0
                        else 0
                    )
                    cell = src_wb[sheet_annexe_1].cell(
                        row=cell.row, column=dico_cols[date_report], value=value
                    )
                else:
                    cell = src_wb[sheet_annexe_1].cell(
                        row=cell.row, column=dico_cols[date_report], value=0
                    )

    data_list = []
    for start, row in enumerate(
        src_wb[sheet_annexe_1].iter_rows(
            min_row=4,
            min_col=column_index_from_string("BL"),
            max_col=column_index_from_string("CU"),
        ),
        start=1,
    ):
        data = []
        for cell in row:
            if has_formula(cell):
                result = interface.calc_cell(cell.coordinate, sheet_annexe_1)
                data.append(result)
            else:
                data.append(cell.value)
        if start == 1:
            new_data = []
            unnamed_counter = 65  # Colonne de début des CMM
            for value in data:
                if value is None:
                    new_data.append(f"Unnamed: {unnamed_counter}")
                    unnamed_counter += 2
                else:
                    new_data.append(value)
            data = new_data
        data_list.append(data)

    df_cmm = pd.concat(
        [df_etat_stock[["code_produit"]], pd.DataFrame(data_list[1:], columns=data_list[0])],
        axis=1,
    )

    df_stock_track_cmm = df_cmm[[col for col in df_cmm.columns if "Unnamed" not in str(col)]]

    df_stock_track_cmm = pd.melt(
        df_stock_track_cmm, id_vars="code_produit", var_name="date_report", value_name="cmm"
    )  # .drop_duplicates()

    df_stock_track_cmm["date_report"] = (
        df_stock_track_cmm["date_report"]
        .apply(lambda x: pd.to_datetime(str(x)[:10], format="%Y-%m-%d"))
        .astype("<M8[ns]")
    )

    interface.clear_cache()
    data_list = []
    for row in src_wb[sheet_annexe_1].iter_rows(
        min_row=3,
        min_col=column_index_from_string("CW"),
        max_col=column_index_from_string("CZ"),
    ):
        data = []
        for cell in row:
            if has_formula(cell):
                result = interface.calc_cell(cell.coordinate, sheet_annexe_1)
                data.append(result)
            else:
                data.append(cell.value)
        data_list.append(data)

    data_list = [row for row in data_list if any(row)]

    df_cmm_currenth_month = pd.concat(
        [df_etat_stock[["code_produit"]], pd.DataFrame(data_list[1:], columns=data_list[0])],
        axis=1,
    ).dropna(how="all")

    df_cmm_currenth_month["date_report"] = pd.to_datetime(date_report, format="%Y-%m-%d")
    df_cmm_currenth_month.columns = df_cmm_currenth_month.columns.str.replace("\n", " ")
    df_cmm_currenth_month.rename(
        columns={
            "Nbre de mois de considérés": "nbre_mois_consideres",
            "Consommations enregistrées sur les mois de considérés": "conso_mois_consideres",
            "CMM Calculée en fin du mois": "cmm_calculee",
            "COMMENTAIRE": "commentaire",
        },
        inplace=True,
    )

    assert (
        df_stock_track_cmm.merge(
            df_cmm_currenth_month, how="left", on=["code_produit", "date_report"]
        ).shape[0]
        == df_stock_track_cmm.shape[0]
    )

    df_stock_track_cmm = df_stock_track_cmm.merge(
        df_cmm_currenth_month, how="left", on=["code_produit", "date_report"]
    )

    df_stock_track_cmm = df_stock_track_cmm.loc[df_stock_track_cmm["date_report"] == date_report]

    cols_df_cmm = df_cmm.columns.to_list()

    mapping = {
        col: cols_df_cmm[i - 1]
        for i, col in enumerate(cols_df_cmm)
        if "Unnamed" in str(col) and i > 0
    }

    df = pd.melt(
        df_cmm,
        id_vars=["code_produit"],
        value_vars=[col for col in df_cmm.columns if not pd.isna(col)],
        var_name="date_report",
    )

    df["date_report"] = df["date_report"].replace(mapping)  # .map(mapping)

    df = df.loc[df.value.notna()]

    df_stock_track_cmm_histo = (
        df.loc[df["value"] == "X"]
        .drop(columns="value")
        .merge(df.loc[df["value"] != "X"], on=["code_produit", "date_report"])
        .rename(columns={"date_report": "date_report_prev"})
        .sort_values(
            ["code_produit", "date_report_prev"], ascending=[True, True], ignore_index=True
        )
    )
    df_stock_track_cmm_histo = df_stock_track_cmm_histo.rename(columns={"value": "cmm"})
    df_stock_track_cmm_histo["date_report"] = pd.to_datetime(date_report, format="%Y-%m-%d")

    return df_stock_track_cmm.round(0), df_stock_track_cmm_histo.round(0)


def get_data_annexe_2(
    src_wb: Workbook,
    sheetnames: list[str],
    df_etat_stock: pd.DataFrame,
    df_stock_prog_nat: pd.DataFrame,
    df_plan_approv: pd.DataFrame,
) -> pd.DataFrame:
    """Extract data from the annex 2 sheet.

    Args:
        src_wb (Workbook): Source workbook containing the annex sheet.
        sheetnames (list[str]): List of sheet names in the workbook.
        df_etat_stock (pd.DataFrame): Data frame containing status stock from sheet annexe 1 - consolidation
        df_stock_prog_nat (pd.DataFrame): DataFrame containing stock program data.
        df_plan_approv (pd.DataFrame): DataFrame containing plan approval data.

    Returns:
        pd.DataFrame: DataFrame containing the data from the annex 2 sheet.
    """
    df_plan_approv["code_and_date_concate"] = df_plan_approv.apply(
        lambda row: str(int(row["Standard product code"])) + "_" + row["DATE"].strftime("%Y-%m-%d")
        if not pd.isna(row["Standard product code"])
        else "_" + row["DATE"].strftime("%Y-%m-%d"),
        axis=1,
    )

    def get_financement_and_delivery_status(code_produit, date_probable_livraison):
        try:
            code_and_date_concate = (
                str(int(code_produit)) + "_" + date_probable_livraison.strftime("%Y-%m-%d")
                if date_probable_livraison
                else np.nan
            )
            index = df_plan_approv.loc[
                df_plan_approv["code_and_date_concate"] == code_and_date_concate
            ].index[0]
            return {
                "financement": df_plan_approv["Source Financement"].iloc[index],
                "status": df_plan_approv["Status"].iloc[index],
            }
        except Exception:
            return {"financement": "", "status": ""}

    sheet_annexe_1, sheet_annexe_2 = (
        check_if_sheet_name_in_file("Annexe 1 - Consolidation", sheetnames),
        check_if_sheet_name_in_file("Annexe 2 - Suivi des Stocks", sheetnames),
    )
    interface = OpenpyxlInterface(wb=src_wb, use_cache=True)
    interface.clear_cache()

    dico_cols = {
        "M": "CONSO",
        "N": "SDU",
        "O": "CMM",
    }

    for start, row in enumerate(
        src_wb[sheet_annexe_2].iter_rows(
            min_row=5,
            max_row=src_wb[sheet_annexe_1].max_row,
            min_col=column_index_from_string("M"),
            max_col=column_index_from_string("O"),
        ),
        start=5,
    ):
        code_produit, facteur_conversion = (
            src_wb[sheet_annexe_1].cell(start, 1).value,
            src_wb[sheet_annexe_1].cell(start, 8).value,
        )
        if not code_produit:
            continue

        df = df_stock_prog_nat.loc[df_stock_prog_nat["Code_produit"] == int(code_produit)]

        # Mise à jour des cellules
        for cell in row:
            if has_formula(cell):
                cell.value = (
                    math.ceil(df[dico_cols.get(cell.column_letter)].sum() / int(facteur_conversion))
                    if not pd.isna(facteur_conversion) and facteur_conversion != 0
                    else 0
                )

        # Mise à jour des éléments de la cellule AJ: "Financement"
        # en supposant que la date probable de livraison n'a pas changé sur base de la date probable de livraison
        cell_financement, cell_date_probable, cell_status, cell_qte_attendue = (
            src_wb[sheet_annexe_2].cell(start, 36),
            src_wb[sheet_annexe_2].cell(start, 37),
            src_wb[sheet_annexe_2].cell(start, 39),
            src_wb[sheet_annexe_2].cell(start, 32),
        )
        if not has_formula(cell_date_probable):
            value = get_financement_and_delivery_status(code_produit, cell_date_probable.value)
            if has_formula(cell_financement):
                cell_financement.value = value["financement"]
            if has_formula(cell_status):
                cell_status.value = value["status"]

        # il faut également faire une correction des formules de la colonne AF: Quantité attendue
        if has_formula(cell_qte_attendue):
            cell_qte_attendue.value = df_plan_approv.loc[
                (df_plan_approv["Standard product code"] == code_produit)
                & (df_plan_approv["DATE"] == cell_date_probable.value),
                "Quantité harmonisée (SAGE)",
            ].sum()

    interface = OpenpyxlInterface(wb=src_wb, use_cache=True)
    interface.clear_cache()
    data_list = []
    columns_letter = []
    for row in src_wb[sheet_annexe_2].iter_rows(
        min_row=5,
        min_col=column_index_from_string("I"),
        max_col=column_index_from_string("AR"),
    ):
        data = []
        for cell in row:
            try:
                # ces colonnes ci ne doivent pas subir de modification en règle générale
                if cell.column_letter in {"AA", "AB", "AC", "AD", "AE"}:
                    continue
                elif has_formula(cell):
                    result = interface.calc_cell(cell.coordinate, sheet_annexe_2)
                    data.append(result)
                else:
                    data.append(cell.value)
            except Exception:
                if cell.column_letter not in columns_letter:
                    columns_letter.append(cell.column_letter)
                # print(e)
                continue
        data_list.append(data)

    return pd.concat(
        [
            df_etat_stock[["code_produit"]],
            pd.DataFrame(data_list, columns=COLUMNS_NAME_ETAT_STOCK),
        ],
        axis=1,
    ).round(0)


def get_data_etat_stock(
    src_wb: Workbook,
    sheetnames: list[str],
    df_etat_stock: pd.DataFrame,
    df_stock_prog_nat: pd.DataFrame,
    df_plan_approv: pd.DataFrame,
    date_report: str,
) -> pd.DataFrame:
    """Extracts and merges stock status data from multiple sources for a given report date.

    Args:
        src_wb (Workbook): The source Excel workbook object.
        sheetnames (list[str]): List of sheet names to extract data from.
        df_etat_stock (pd.DataFrame): DataFrame containing the current stock status.
        df_stock_prog_nat (pd.DataFrame): DataFrame containing national stock program data.
        df_plan_approv (pd.DataFrame): DataFrame containing approval plan data.
        date_report (str): The report date in 'YYYY-MM-DD' format.

    Returns:
        pd.DataFrame: The merged and processed stock status DataFrame with updated columns and report date.

    Raises:
        AssertionError: If the merge operation does not preserve the number of rows in df_etat_stock.
    """
    df_data = get_data_annexe_2(
        src_wb=src_wb,
        sheetnames=sheetnames,
        df_etat_stock=df_etat_stock,
        df_stock_prog_nat=df_stock_prog_nat,
        df_plan_approv=df_plan_approv.copy(),
    )
    assert (
        df_etat_stock.merge(df_data, on="code_produit", how="inner").shape[0]
        == df_etat_stock.shape[0]
    )

    df_etat_stock = df_etat_stock.merge(df_data, on="code_produit", how="inner")
    df_etat_stock = df_etat_stock.rename(columns=DICO_COLUMNS)
    df_etat_stock["date_report"] = pd.to_datetime(date_report, format="%Y-%m-%d")

    return df_etat_stock.round(0)
